"""
test_excel_calc.py — เทส excel_calc.py (คำนวณสูตรจริงด้วย LibreOffice)

เทสพวกนี้ต้องมี LibreOffice ติดตั้งอยู่ ถ้าไม่มีจะ skip เอง (ไม่ทำให้ test suite พัง) — สร้างไฟล์
.xlsx ที่มีสูตรขึ้นเองในเทส ไม่พึ่งไฟล์เทมเพลตจริงของลูกค้า
"""
import os
import shutil
import tempfile
import unittest

import openpyxl
from django.test import SimpleTestCase

from customers.cpall.logic.excel_calc import ExcelCalcError, load_calculated_workbook

HAS_LIBREOFFICE = shutil.which("soffice") is not None


def _make_xlsx_with_formulas():
    """สร้างไฟล์ทดสอบที่มีสูตรแบบเดียวกับที่เทมเพลตจริงใช้ (SUM, SUMPRODUCT+ROUNDUP)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "test"
    ws["A1"], ws["B1"], ws["C1"] = 10, 20, 30
    ws["D1"] = "=SUM(A1:C1)"              # -> 60
    ws["A2"] = 36                          # pack size
    ws["D2"] = "=SUMPRODUCT(ROUNDUP(A1:C1/A2, 0))"  # ceil(10/36)+ceil(20/36)+ceil(30/36) = 3
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


@unittest.skipUnless(HAS_LIBREOFFICE, "ต้องมี LibreOffice ติดตั้งอยู่ถึงจะรันเทสนี้ได้")
class ExcelCalcTests(SimpleTestCase):
    def test_calculates_sum_formula(self):
        path = _make_xlsx_with_formulas()
        try:
            wb = load_calculated_workbook(path)
            self.assertEqual(wb["test"]["D1"].value, 60)
        finally:
            os.remove(path)

    def test_calculates_sumproduct_roundup_formula(self):
        """สูตรเดียวกับที่ใช้คำนวณ 'รวมตะกร้า' ในเทมเพลตจริง"""
        path = _make_xlsx_with_formulas()
        try:
            wb = load_calculated_workbook(path)
            self.assertEqual(wb["test"]["D2"].value, 3)
        finally:
            os.remove(path)

    def test_original_file_not_modified(self):
        """ไฟล์ต้นฉบับต้องไม่ถูกแก้ (สูตรยังอยู่ครบ ไม่ถูกแทนที่ด้วยค่า)"""
        path = _make_xlsx_with_formulas()
        try:
            load_calculated_workbook(path)
            wb_orig = openpyxl.load_workbook(path)  # data_only=False = อ่านสูตรดิบ
            self.assertEqual(wb_orig["test"]["D1"].value, "=SUM(A1:C1)")
        finally:
            os.remove(path)

    def test_missing_file_raises(self):
        with self.assertRaises(ExcelCalcError):
            load_calculated_workbook("/tmp/ไฟล์ที่ไม่มีอยู่จริง.xlsx")
