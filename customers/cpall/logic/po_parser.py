"""
po_parser.py — Module 1: PO Parser + Loader

อ่านไฟล์ PO Export (.xlsx) รูปแบบเดียวกับ cpall_po_export_20260825.xlsx
ตรวจสอบว่ามีคอลัมน์ที่จำเป็นครบ แล้วบันทึกลง Postgres (po_import + po_line)

*** Phase 1 (ORM migration) *** ฟังก์ชัน CRUD ทั้งหมดในไฟล์นี้ย้ายมาใช้ Django ORM แล้ว (เดิมเป็น
raw SQL) — parse_po_file() ไม่เปลี่ยน (ยังใช้ pandas อ่าน Excel เหมือนเดิม ไม่เกี่ยวกับ DB) ข้อดีที่ได้
เพิ่มมาโดยไม่ได้ตั้งใจ: load_po_to_db() เดิม INSERT ทีละแถว (300+ ครั้งต่อไฟล์) ตอนนี้ใช้
PoLine.objects.bulk_create() แทน เร็วขึ้นจริงเพราะยิง query เดียวหลายแถวแทนที่จะยิงทีละแถว
"""
import os
from datetime import date, datetime

import openpyxl
import pandas as pd

from customers.cpall.logic.db import get_cpall_customer_id
from customers.cpall.models import LocationMapping, PlanRun, PoImport, PoLine, ProductMaster

# คอลัมน์ที่ต้องมีใน PO Export — ถ้าไม่ครบ ให้หยุดทันที (ตาม FR-1 / UC-1 exception)
REQUIRED_COLUMNS = [
    "Purchase Order Number",
    "Purchase Order Date",
    "Delivery Date",
    "Delivery Time",
    "Delivery Location Number",
    "Delivery Location",
    "Line Item Number",
    "Item Number (Product Code)",
    "Item Name ",          # หมายเหตุ: ไฟล์ต้นฉบับมีช่องว่างท้ายชื่อคอลัมน์นี้จริง
    "Ordered Quantity",
    "Unit Type ",          # เช่นกัน มีช่องว่างท้าย
    "Net Case Price",
]


class POParseError(Exception):
    pass


class POInUseError(Exception):
    """ลบไม่ได้เพราะ po_import นี้ถูกใช้สร้างแผนไปแล้ว"""
    pass


def _json_safe(value):
    """
    แปลงค่าจาก pandas/openpyxl ให้เก็บลง JSONField ได้ — ค่าที่ JSON เข้ารหัสตรงๆ ไม่ได้มี pd.NaT/NaN
    (float พิเศษ, ไม่ใช่ None) และ datetime/date/Timestamp (ไม่ใช่ string) ต้องแปลงก่อนเสมอ

    ไฟล์ PO ต้นฉบับปกติเก็บวันที่เป็น "ข้อความ" dd/mm/yyyy อยู่แล้ว (ไม่ใช่ Excel date object) แต่ถ้า
    บางไฟล์ดันเป็น Excel date object จริง (pandas parse เป็น datetime/Timestamp อัตโนมัติ) ต้อง format
    เป็น dd/mm/yyyy ให้ตรงกับที่ไฟล์ต้นฉบับแบบข้อความใช้ — ไม่ใช้ .isoformat() (yyyy-mm-dd) เพราะจะทำให้
    ไฟล์ที่สร้างใหม่ (ตอนต้นฉบับหายไปแล้ว) แสดงผลวันที่ต่างจากต้นฉบับที่ Admin คุ้นเคย
    """
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    if hasattr(value, "strftime"):  # pandas.Timestamp และอื่นๆ ที่มีเมธอดนี้แต่ไม่ใช่ subclass ของ datetime
        return value.strftime("%d/%m/%Y")
    return value


def _to_date(value):
    """แปลงวันที่รูปแบบ 'DD/MM/YYYY' (ปี พ.ศ. หรือ ค.ศ. ตามที่ระบบลูกค้าส่งมา) เป็น datetime.date"""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None  # ไม่ parse ได้ -> เก็บเป็น NULL ดีกว่าทำให้ทั้งแถวพัง


def parse_po_file(filepath: str) -> pd.DataFrame:
    """
    อ่านไฟล์ PO Export คืนค่าเป็น DataFrame ที่ normalize คอลัมน์แล้ว
    เก็บ "ค่าทุกคอลัมน์ตามลำดับเดิม" ไว้ในคอลัมน์พิเศษ "_all_values" ของแต่ละแถวด้วย (list) และ
    "ลำดับ+ชื่อคอลัมน์ต้นฉบับ" ไว้ใน out.attrs["column_order"] — ใช้ตอนสร้างไฟล์ใหม่ที่มีข้อมูลครบ
    เหมือนต้นฉบับทีหลัง (ดู po_regenerator.py) แม้ว่า REQUIRED_COLUMNS จะใช้แค่ 12 คอลัมน์จาก 50+
    คอลัมน์ที่ไฟล์จริงมีก็ตาม — เก็บด้วย "ตำแหน่ง" ไม่ใช่ "ชื่อ" เพราะไฟล์จริงมีชื่อคอลัมน์ซ้ำกันได้
    (เช่น "Discount Percentage 1" ปรากฏ 2 รอบ) อ่าน header ต้นฉบับผ่าน openpyxl ตรงๆ แทนที่จะใช้
    df.columns ของ pandas เพราะ pandas จะ auto-rename ชื่อซ้ำเป็น ".1"/".2" ให้เอง (ไม่ใช่ชื่อจริง)
    """
    wb_raw = openpyxl.load_workbook(filepath, data_only=True)
    ws_raw = wb_raw.active
    original_column_order = [c.value for c in next(ws_raw.iter_rows(min_row=1, max_row=1))]
    # เก็บค่าดิบทุกแถวข้อมูล (ไม่ผ่าน pandas) ไว้คู่กับ column_order — อ่านตรงจาก openpyxl รักษา
    # type ดั้งเดิมของแต่ละ cell แม่นยำกว่า pandas (pandas ทำ type inference ให้อัตโนมัติแม้ตั้ง
    # dtype=object ไว้แล้วก็ตาม เช่น เลขจำนวนเต็มในไฟล์อาจถูกแปลงเป็น float โดยไม่ตั้งใจ)
    raw_rows_by_excel_index = {
        i: list(row) for i, row in enumerate(ws_raw.iter_rows(min_row=2, values_only=True))
    }
    wb_raw.close()

    df = pd.read_excel(filepath, sheet_name=0, dtype=object)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise POParseError(
            f"ไฟล์ PO ขาดคอลัมน์ที่จำเป็น: {missing}\n"
            f"คอลัมน์ที่พบในไฟล์: {list(df.columns)}"
        )

    out = pd.DataFrame()
    out["po_number"] = df["Purchase Order Number"].astype(str).str.strip()
    out["po_date"] = df["Purchase Order Date"].apply(_to_date)
    out["delivery_date"] = df["Delivery Date"].apply(_to_date)
    out["delivery_time"] = df["Delivery Time"].astype(str)
    out["fc_code"] = df["Delivery Location Number"].astype(str).str.strip()
    out["delivery_location"] = df["Delivery Location"]
    out["line_no"] = pd.to_numeric(df["Line Item Number"], errors="coerce")
    out["barcode"] = df["Item Number (Product Code)"].astype(str).str.strip()
    out["item_name"] = df["Item Name "]
    out["qty_ordered"] = pd.to_numeric(df["Ordered Quantity"], errors="coerce")
    out["unit_type"] = df["Unit Type "]
    out["net_case_price"] = pd.to_numeric(df["Net Case Price"], errors="coerce")
    # เก็บค่าทุกคอลัมน์ตามตำแหน่งเดิม จากค่าดิบที่อ่านผ่าน openpyxl (ไม่ใช่ pandas — รักษา
    # type ดั้งเดิมแม่นยำกว่า) จับคู่ด้วย df.index ที่ตรงกับลำดับแถวดิบเสมอ (pandas ไม่ reset index
    # ตอน filter ทีหลัง จึงยังจับคู่ย้อนกลับไปที่แถวดิบถูกต้องแม้หลัง dropna/dedup)
    out["_all_values"] = pd.Series(
        [[_json_safe(v) for v in raw_rows_by_excel_index.get(i, [])] for i in df.index],
        index=df.index, dtype=object,
    )

    # แถวที่ไม่มี PO number หรือ barcode คือแถวว่าง/สรุปท้ายไฟล์ -> ตัดทิ้ง
    out = out.dropna(subset=["po_number", "barcode"])
    out = out[out["po_number"].str.strip() != ""]

    # เช็คแถวที่ซ้ำกันเป๊ะภายในไฟล์เดียวกัน (PO เดียวกัน + จุดส่งเดียวกัน + SKU เดียวกัน + line_no
    # เดียวกัน) — เจอได้บางครั้งจากไฟล์ export ต้นทางเอง ไม่ใช่ความผิดของเรา แต่ถ้าไม่กรองออกจะทำให้
    # ยอดสั่งเพี้ยน (นับซ้ำ) — ตัดแถวซ้ำออก เก็บแถวแรกที่เจอไว้ แล้วแจ้งจำนวนที่ตัดออกให้เห็นชัดเจน
    dup_key_cols = ["po_number", "fc_code", "barcode", "line_no"]
    dup_mask = out.duplicated(subset=dup_key_cols, keep="first")
    n_dup = int(dup_mask.sum())
    if n_dup > 0:
        print(f"[po_parser] ⚠️  พบแถวข้อมูลซ้ำกันเป๊ะ {n_dup} แถว (PO+จุดส่ง+SKU+ลำดับเดียวกัน) — ตัดออกอัตโนมัติ")
        out = out[~dup_mask]

    # เตือน (ไม่ raise) ถ้า unit_type ไม่ใช่ CT ตามที่ระบุไว้ใน Data Model (ข้อ 2.4)
    non_ct = out[~out["unit_type"].isin(["CT"])]
    if len(non_ct) > 0:
        print(f"[po_parser] ⚠️  พบ {len(non_ct)} แถวที่ Unit Type ไม่ใช่ 'CT' — ควรตรวจสอบ")

    out.attrs["column_order"] = original_column_order
    return out


def load_po_to_db(filepath: str, production_date, po_date, imported_by: str = "admin") -> int:
    """
    parse ไฟล์ PO แล้วบันทึกลง Postgres คืนค่า po_import_id ที่สร้างขึ้น
    production_date, po_date: datetime.date — วันที่ผลิต / วันที่ PO ของรอบนี้ (ผูกกับรอบนี้ตั้งแต่ตอน
    import เพราะ Production Plan รวมหลายรอบเข้าไฟล์เดียว แต่ละรอบอาจมีวันที่ต่างกัน — เช่น รอบบ่ายผลิต
    วันนี้ส่งพรุ่งนี้ แต่รอบเช้าต่างจังหวัดของวันถัดมาอาจวันที่ต่างออกไป)
    """
    df = parse_po_file(filepath)
    customer_id = get_cpall_customer_id()
    column_order = df.attrs.get("column_order")

    po_import = PoImport.objects.create(
        customer_id=customer_id, source_filename=filepath, imported_by=imported_by,
        production_date=production_date, po_date=po_date, total_rows=len(df), status="imported",
        column_order=column_order,
    )

    lines = []
    for _, row in df.iterrows():
        qty = float(row["qty_ordered"]) if pd.notna(row["qty_ordered"]) else None
        price = float(row["net_case_price"]) if pd.notna(row["net_case_price"]) else None
        lines.append(PoLine(
            po_import=po_import,
            po_number=row["po_number"],
            po_date=row["po_date"],
            delivery_date=row["delivery_date"],
            delivery_time=row["delivery_time"],
            fc_code=row["fc_code"],
            delivery_location=row["delivery_location"],
            line_no=int(row["line_no"]) if pd.notna(row["line_no"]) else None,
            barcode=row["barcode"],
            item_name=row["item_name"],
            qty_ordered=qty,
            unit_type=row["unit_type"],
            net_case_price=price,
            total_amount=(qty * price) if qty is not None and price is not None else None,
            all_values=row["_all_values"],
        ))
    PoLine.objects.bulk_create(lines)

    print(f"[po_parser] imported {len(df)} lines from {filepath} -> po_import_id={po_import.id}")
    return po_import.id


def delete_po_import(po_import_id: int):
    """
    ลบ PO ที่ import ไว้ (แถว po_line + po_import + ไฟล์ที่อัปโหลดไว้บนดิสก์ ถ้ายังมีอยู่จริง — PO
    ใหม่ตั้งแต่มีระบบ data-first ไม่มีไฟล์ค้างอยู่แล้วปกติ มีแค่ PO เก่าก่อนหน้าที่อาจยังมีไฟล์อยู่)
    ถ้าเคยถูกใช้สร้างแผนไปแล้ว (มีอยู่ใน plan_run_import) จะ raise POInUseError ทันที ไม่ลบอะไรเลย
    — ต้องไปลบแผนที่ใช้ PO นี้ก่อน ถึงจะลบ PO นี้ได้ (ไม่มีปุ่ม "บังคับลบ" ข้ามการเช็คนี้แล้ว —
    เดิมมีไว้เผื่อไฟล์ PO หายไปจากที่เก็บ แต่ตอนนี้ PO เป็น data-first เต็มรูปแบบแล้ว ไม่มีไฟล์ให้หายอีก)
    ถ้า PO นี้ถูกลบไปแล้ว (หาไม่เจอ) ถือว่าลบสำเร็จเงียบๆ ไม่ error — กันเคส race condition/กดซ้ำ
    """
    used_plan_runs = PlanRun.objects.filter(planrunimport__po_import_id=po_import_id).distinct()
    if used_plan_runs.exists():
        plan_labels = ", ".join(p.get_short_label() for p in used_plan_runs)
        raise POInUseError(
            f"PO นี้ถูกใช้สร้างแผนไปแล้ว ({plan_labels}) — ลบแผนเหล่านี้ก่อน ถึงจะลบ PO นี้ได้"
        )

    try:
        po_import = PoImport.objects.get(id=po_import_id)
    except PoImport.DoesNotExist:
        return

    source_filename = po_import.source_filename
    po_import.delete()  # PoLine ตั้ง on_delete=CASCADE ไว้แล้ว ลบตามอัตโนมัติ ไม่ต้อง DELETE แยก

    if source_filename and os.path.exists(source_filename):
        os.remove(source_filename)


def list_po_imports_paginated(page: int = 1, page_size: int = 10, search: str = "") -> dict:
    """
    ดึงรายการ PO แบบแบ่งหน้า + ค้นหาได้ — ใช้ที่หน้า "PO ทั้งหมด" (ต่างจาก list_po_imports() เดิม
    ที่ใช้แค่โชว์ตัวอย่างสั้นๆ ในแดชบอร์ด) ค้นหา/แบ่งหน้าที่ระดับ SQL ตรงๆ ผ่าน ORM (ไม่ใช่ดึงมาทั้งหมด
    แล้วตัดทีหลัง) — รองรับข้อมูลเยอะในอนาคตได้โดยไม่ช้าลง
    -> {"items": [...], "total": N, "page": ..., "page_size": ..., "total_pages": ...}
    """
    page = max(1, page)
    page_size = page_size if page_size in (10, 50, 100) else 10
    offset = (page - 1) * page_size

    qs = PoImport.objects.all()
    if search:
        qs = qs.filter(source_filename__icontains=search)

    total = qs.count()
    rows = qs.order_by("-imported_at")[offset:offset + page_size]

    items = [
        {"id": r.id, "source_filename": r.source_filename,
         "display_filename": os.path.basename(r.source_filename), "production_date": r.production_date,
         "po_date": r.po_date, "total_rows": r.total_rows,
         "imported_at": r.imported_at, "imported_by": r.imported_by}
        for r in rows
    ]
    total_pages = max(1, (total + page_size - 1) // page_size)
    return {"items": items, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}


def list_po_imports(limit: int = 50) -> list[dict]:
    """
    ดึงรายการ PO ที่เคย import ไว้ทั้งหมด เรียงล่าสุดก่อน — ใช้แสดงในหน้า Dashboard
    -> [{"id":, "source_filename":, "production_date":, "po_date":, "total_rows":, "imported_at":, "imported_by":}, ...]
    """
    rows = PoImport.objects.order_by("-imported_at")[:limit]
    return [
        {"id": r.id, "source_filename": r.source_filename,
         "display_filename": os.path.basename(r.source_filename), "production_date": r.production_date,
         "po_date": r.po_date, "total_rows": r.total_rows,
         "imported_at": r.imported_at, "imported_by": r.imported_by}
        for r in rows
    ]


def check_unknown_locations(po_import_id: int) -> list:
    """UC-2: หา fc_code ในรอบนี้ที่ยังไม่มีใน location_mapping"""
    known_fc_codes = LocationMapping.objects.values_list("fc_code", flat=True)
    rows = (
        PoLine.objects.filter(po_import_id=po_import_id)
        .exclude(fc_code__in=known_fc_codes)
        .values_list("fc_code", "delivery_location")
        .distinct()
    )
    return list(rows)


def check_unknown_skus(po_import_id: int) -> list:
    """
    หาบาร์โค้ดในรอบนี้ที่ยังไม่มีใน product_master (เหมือน check_unknown_locations แต่สำหรับสินค้า) —
    ไม่มีผลต่อการคำนวณแผนเลย (product_master ใช้แค่แสดงชื่อสินค้าที่หน้ากรอกยอดเผื่อ) แต่ยังอยากให้
    Admin กรอกไว้ให้ครบเพื่อความสมบูรณ์ของข้อมูลอ้างอิง — คืน (barcode, item_name, net_case_price)
    ต่อบาร์โค้ดที่ไม่รู้จัก — ราคาเอาไว้ auto-fill ในฟอร์ม (มีอยู่แล้วในไฟล์ PO ไม่ต้องให้กรอกซ้ำ)
    """
    known_barcodes = ProductMaster.objects.values_list("barcode", flat=True)
    rows = (
        PoLine.objects.filter(po_import_id=po_import_id)
        .exclude(barcode__in=known_barcodes)
        .order_by("barcode")
        .values_list("barcode", "item_name", "net_case_price")
        .distinct("barcode")  # บาร์โค้ดเดียวกันอาจมีหลายแถว (คนละจุดส่ง) เอาแค่ตัวแทนตัวเดียวพอ
    )
    return list(rows)
