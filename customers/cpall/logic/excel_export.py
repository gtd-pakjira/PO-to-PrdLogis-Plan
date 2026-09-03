"""
excel_export.py — Module 4: Excel Exporter (Production Plan)

เขียนทับ Template `cpall/excel_templates/production_plan_template.xlsx` (แพลน 7-11.xlsx เดิม)
ด้วยยอดที่คำนวณจาก Module 2 (ระดับ sub_location) โดยคงหัวคอลัมน์/รูปแบบเดิมไว้ทั้งหมด

*** อัปเดต: ไม่ hardcode ตำแหน่งคอลัมน์อีกต่อไป ***
โค้ดจะสแกนแถว 7-8 ของเทมเพลตเอง เพื่อหาว่าคอลัมน์ไหนคือจุดส่งย่อยอะไร (บางบัวทอง/ชลบุรี/.../หาดใหญ่)
แล้วกรอกยอดแยกทีละคอลัมน์ตามจุดส่งย่อยจริง (ไม่ใช่แค่ยอดรวม + 2 กลุ่มที่ตรง 1:1 แบบเดิม)
ถ้าเทมเพลตเปลี่ยนตำแหน่งคอลัมน์ในอนาคต โค้ดนี้ยังทำงานถูก เพราะอ่านตำแหน่งจากหัวตารางเอง ไม่ใช่เลขคอลัมน์ตายตัว

วิธีรัน (จาก root ของโปรเจกต์):
    python -m src.excel_export <po_import_id> <output_path>
"""
import re
import sys

import openpyxl

from customers.cpall.logic.date_utils import update_date_headers
from customers.cpall.logic.grouping import get_grouped_quantities_by_sub_location

TEMPLATE_PATH = "customers/cpall/excel_templates/production_plan_template.xlsx"
SHEET_NAME = "แพลน 7-11"


class ExcelExportError(Exception):
    pass


COL_LINE_NO = 3       # C — Line no. (ใช้หาแถวหัว SKU)
COL_NAME = 4          # D
BARCODE_ROW_OFFSET = 2  # แถวบาร์โค้ด = แถวหัว SKU + 2
BUFFER_ROW_OFFSET = 2   # แถวยอดเผื่อ = แถวเดียวกับแถวบาร์โค้ด (header_row + 2)
BUFFER_COL = 15         # คอลัมน์ O — ตรวจสอบกับไฟล์จริงแล้วว่ายอดเผื่ออยู่คอลัมน์นี้เสมอทุก SKU
                         # (ไม่มีหัวตารางกำกับคอลัมน์นี้ในเทมเพลต เลย hardcode ไว้ตรงนี้ — ถ้าเทมเพลต
                         # เปลี่ยนผังในอนาคต ต้องมาแก้ค่านี้)

HEADER_GROUP_ROW = 7       # แถวที่มีชื่อกลุ่ม/สถานที่หลัก เช่น "บางบัวทอง", "มหาชัย"
HEADER_SUB_ROW = 8         # แถวที่มีชื่อจุดส่งย่อย เช่น "ชลบุรี", "ขอนแก่น"
HEADER_TOTAL_LABEL_ROW = 5 # แถวที่มีคำว่า "ยอดรวม"
HEADER_SCAN_COL_RANGE = range(5, 25)  # ช่วงคอลัมน์ที่สแกนหาหัวตาราง (กันเผื่อเทมเพลตขยับ)

# ป้ายที่ไม่ใช่ชื่อจุดส่งย่อยจริง (ข้อความหมายเหตุที่แทรกอยู่ในแถวหัวกลุ่มของเทมเพลต) — ข้ามไปถ้าเจอ
IGNORE_HEADER_LABELS = {"บาร์ระบุวันผลิต"}

# เทมเพลตเดิมพิมพ์ชื่อจุดส่งย่อยไม่ตรงกับที่ตั้งไว้ใน location_mapping.yaml (ย่อ/พิมพ์ตก) — แก้ให้ตรงกันตรงนี้
SUB_LOCATION_LABEL_CORRECTIONS = {
    "สุราษร์": "สุราษฎร์ธานี",
    "นครราขสีมา": "นครราชสีมา",  # เทมเพลตสะกดผิด (ข แทน ช)
}

# บาร์โค้ดที่เทมเพลตพิมพ์ผิด/สลับกับ SKU อื่น (พบระหว่างสร้าง sku_master.yaml)
LEGACY_TEMPLATE_BARCODE_CORRECTIONS = {
    "88859388000063": "8859388000063",   # องุ่นเคียวโฮ 300 กรัม
    "88593988001091": "8859388001091",   # ลูกไหนจิ๋ว 200 กรัม
}


def _find_sub_location_columns(ws) -> dict:
    """
    สแกนแถว HEADER_GROUP_ROW และ HEADER_SUB_ROW เพื่อหา (ชื่อจุดส่งย่อย -> เลขคอลัมน์)
    กติกา: ถ้าแถวย่อย (8) มีข้อความ ใช้อันนั้นก่อน (เจาะจงกว่า) ถ้าไม่มี ค่อยใช้แถวกลุ่ม (7)
    """
    col_to_sub_location = {}
    for col in HEADER_SCAN_COL_RANGE:
        sub_label = ws.cell(row=HEADER_SUB_ROW, column=col).value
        group_label = ws.cell(row=HEADER_GROUP_ROW, column=col).value

        label = None
        if sub_label and str(sub_label).strip():
            label = str(sub_label).strip()
        elif group_label and str(group_label).strip() not in IGNORE_HEADER_LABELS:
            label = str(group_label).strip()

        if label:
            label = SUB_LOCATION_LABEL_CORRECTIONS.get(label, label)
            col_to_sub_location[col] = label

    return col_to_sub_location


def _find_total_column(ws):
    """สแกนหาคอลัมน์ที่หัวตารางเขียนว่า 'ยอดรวม' (ไม่ hardcode เลขคอลัมน์)"""
    for col in HEADER_SCAN_COL_RANGE:
        val = ws.cell(row=HEADER_TOTAL_LABEL_ROW, column=col).value
        if val and "ยอดรวม" in str(val):
            return col
    return None


def _find_sku_header_rows(ws) -> dict:
    """
    สแกนชีตหา (barcode -> แถวหัว SKU) โดยไม่ hardcode เลขแถว
    (แถวหัว SKU = แถวที่คอลัมน์ C เป็นตัวเลข/ลำดับ, บาร์โค้ดอยู่ 2 แถวถัดไปในคอลัมน์ D)
    """
    mapping = {}
    for row in range(1, ws.max_row + 1):
        line_no = ws.cell(row=row, column=COL_LINE_NO).value
        if not isinstance(line_no, (int, float)):
            continue

        header_name = str(ws.cell(row=row, column=COL_NAME).value or "")
        barcode_cell = ws.cell(row=row + BARCODE_ROW_OFFSET, column=COL_NAME).value
        if not barcode_cell:
            continue

        m = re.match(r"\s*(\d{10,14})", str(barcode_cell))
        if not m:
            continue
        raw_barcode = m.group(1)

        # เคสพิเศษ: บาร์โค้ดในเทมเพลตเป็นเลขจริงของ SKU อื่น (ไม่ใช่แค่พิมพ์ผิด) -> เช็คจากชื่อ SKU แทน
        if "องุ่นดำไร้เมล็ด 300" in header_name and raw_barcode == "8859388010284":
            mapping["8859388002562"] = row
            continue

        corrected = LEGACY_TEMPLATE_BARCODE_CORRECTIONS.get(raw_barcode, raw_barcode)
        mapping[corrected] = row

    return mapping


def read_buffer_qty_from_logistic_plan(filepath: str, sheet_name: str = "บางบัวทอง-ผลิต") -> dict:
    """
    อ่านค่า "ยอดเผื่อ" ต่อ SKU จากไฟล์ Logistic Plan ที่ Admin กรอกไว้แล้ว (เช่นไฟล์รอบเช้าต่างจังหวัด
    ที่เพิ่งสร้างแล้วเปิดไปกรอกยอดเผื่อเพิ่ม) — หาคอลัมน์ 'ยอดเผื่อ' เองแบบไดนามิก ไม่ hardcode
    -> {barcode: buffer_qty}

    หมายเหตุ: Phase 1 ยังไม่รู้สูตรคำนวณยอดเผื่อ (Admin ยังกรอกเองอยู่) — ฟังก์ชันนี้แค่ "อ่านค่าที่มีอยู่แล้ว"
    มาใช้ต่อ ไม่ได้คำนวณเอง เมื่อรู้สูตรแล้วค่อยเปลี่ยนมาคำนวณอัตโนมัติแทนการอ่านจากไฟล์
    """
    import re as _re
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    buffer_col = None
    for row in range(1, 15):
        for col in range(1, 30):
            val = ws.cell(row=row, column=col).value
            if val and "ยอดเผื่อ" in str(val):
                buffer_col = col
                break
        if buffer_col:
            break
    if buffer_col is None:
        return {}

    result = {}
    for row in range(1, ws.max_row + 1):
        line_no = ws.cell(row=row, column=2).value
        # หาคอลัมน์ชื่อ/บาร์โค้ดแบบเดียวกับ logistic_plan_export (line_no อยู่ก่อนชื่อ 1 คอลัมน์เสมอ)
        if not isinstance(line_no, (int, float)):
            # ไฟล์บางไฟล์ line_no อยู่ col 3 (เช่น สุวรรณภูมิ/รอบเช้าต่างจังหวัด) ลองเผื่อไว้
            line_no = ws.cell(row=row, column=3).value
            name_col = 4
        else:
            name_col = 3
        if not isinstance(line_no, (int, float)):
            continue

        barcode_cell = ws.cell(row=row + 1, column=name_col).value
        if not barcode_cell:
            continue
        m = _re.search(r"\d{10,16}", str(barcode_cell))
        if not m:
            continue
        barcode = m.group(0)

        buffer_val = ws.cell(row=row, column=buffer_col).value
        if buffer_val is not None:
            result[barcode] = buffer_val

    return result


def write_buffer_qty(ws, header_rows: dict, buffer_by_barcode: dict, search_col_range=range(5, 20)):
    """
    เขียนค่า "ยอดเผื่อ" ลง Production Plan ตรงแถว (ยอดเผื่อ) ของแต่ละ SKU
    หาคอลัมน์ที่ต้องเขียนแบบไดนามิก: สแกนแถว (header_row + 2) หาช่องที่เป็นตัวเลข (ไม่ใช่ช่อง
    บาร์โค้ด/ชื่อ) — ถ้าไม่เจอเลย (แถวว่างสนิท) ใช้คอลัมน์เดียวกับที่เจอในไฟล์ตัวอย่างจริงเป็นค่าเริ่มต้น (O)
    คืนค่าจำนวน SKU ที่เขียนสำเร็จ
    """
    FALLBACK_COL = 15  # คอลัมน์ O — ตำแหน่งที่พบจริงในไฟล์ Production Plan ตัวอย่าง (สแกนแล้ว 7 SKU ตรงหมด)
    written = 0

    for barcode, buffer_val in buffer_by_barcode.items():
        header_row = header_rows.get(barcode)
        if header_row is None:
            continue
        buffer_row = header_row + 2

        target_col = None
        for col in search_col_range:
            val = ws.cell(row=buffer_row, column=col).value
            if isinstance(val, (int, float)):
                target_col = col
                break
        if target_col is None:
            target_col = FALLBACK_COL

        ws.cell(row=buffer_row, column=target_col).value = buffer_val
        written += 1

    return written


def export_production_plan(po_import_ids, output_path: str, buffer_override: dict = None,
                            template_path: str = TEMPLATE_PATH):
    """
    po_import_ids: รับได้ทั้ง int เดี่ยว หรือ list ของ int (รวมหลายรอบ PO เข้าด้วยกัน)
    เช่น Production Plan ที่ต้องรวมรอบบ่าย (บางบัวทอง/มหาชัย/สุวรรณภูมิ) กับรอบเช้า (รอบเช้าต่างจังหวัด)
    เข้าด้วยกันเป็นแผนเดียว — ถ้าส่งมาแค่รอบเดียว (เพราะอีกรอบยังไม่มา) ก็ยังทำงานได้ปกติ
    เพียงแต่จุดส่งที่ยังไม่มีข้อมูลจะเว้นว่างไว้ (ไม่ error)

    วันที่ในหัวไฟล์: ดึงจากวันที่ที่ผูกไว้กับแต่ละ po_import_id ตอน import (ไม่ต้องส่งเข้ามาซ้ำที่นี่)
    เพราะ Production Plan รวมหลายรอบที่อาจมีวันที่ต่างกัน (เช่น รอบบ่ายส่งพรุ่งนี้ รอบเช้าต่างจังหวัด
    ส่งวันเดียวกับที่ PO เข้า) — แต่ละคอลัมน์จะได้วันที่ของรอบที่ตัวเองสังกัดเอง

    buffer_override: {barcode: qty} — ถ้าระบุมา จะใช้ค่านี้แทนการอ่านจากไฟล์เทมเพลต "รอบเช้าต่างจังหวัด"
    (เช่น ตอน Admin กรอกยอดเผื่อผ่านหน้าเว็บเอง — ดู UC-4) ถ้าไม่ระบุ (None) จะ fallback ไปอ่านจากเทมเพลต
    เหมือนเดิม (behavior เดิมที่จำลอง Admin กรอกไว้ในไฟล์)
    """
    from customers.cpall.logic.grouping import get_covered_sub_locations, get_dates_by_sub_location
    sub_location_qty = get_grouped_quantities_by_sub_location(po_import_ids)
    covered_sub_locations = get_covered_sub_locations(po_import_ids)
    dates_by_sub_location = get_dates_by_sub_location(po_import_ids)

    # จัดรูปเป็น {barcode: {sub_location: qty}}
    qty_by_barcode = {}
    for row in sub_location_qty:
        qty_by_barcode.setdefault(row["barcode"], {})[row["sub_location"]] = row["qty_case_ordered"]

    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET_NAME]

    col_to_sub_location = _find_sub_location_columns(ws)

    def date_resolver(col):
        """หาว่าคอลัมน์นี้ (หัว 'วันที่...') สังกัดจุดส่งย่อยไหน แล้วคืนวันที่ของรอบที่จุดนั้นสังกัด
        (เซลล์หัว 'วันที่' อยู่คอลัมน์เดียวกับจุดส่งย่อยแรกใต้หัวนั้นเสมอ เพราะเป็น merged cell)"""
        sub_loc = col_to_sub_location.get(col)
        if sub_loc is None or sub_loc not in dates_by_sub_location:
            return None
        production_date, po_date = dates_by_sub_location[sub_loc]
        if production_date is None or po_date is None:
            return None
        return production_date, po_date

    n = update_date_headers(ws, date_resolver)
    print(f"[excel_export] อัปเดตวันที่ในหัวไฟล์ {n} จุด (แต่ละจุดใช้วันที่ของรอบ PO ที่ตัวเองสังกัด)")

    total_col = _find_total_column(ws)
    header_rows = _find_sku_header_rows(ws)

    print(f"[excel_export] อ่านหัวตารางได้ {len(col_to_sub_location)} คอลัมน์จุดส่งย่อย: "
          f"{list(col_to_sub_location.values())}")
    print(f"[excel_export] คอลัมน์ 'ยอดรวม' อยู่ที่คอลัมน์ {total_col} — มีสูตร SUM อยู่แล้วในเทมเพลต "
          f"จะไม่เขียนทับ (ปล่อยให้ Excel คำนวณเองจากค่าที่กรอกใน G:Q)")

    filled, missing_in_template, unmatched_sub_locations = [], [], set()

    # ---------- ยอดเผื่อ: ใช้ค่าที่ Admin กรอกผ่านเว็บ (ถ้ามี) ไม่งั้น fallback ไปอ่านจากไฟล์เทมเพลต ----------
    # ดูหมายเหตุใน src/logistic_plan_export.read_buffer_qty_from_template — การอ่านจากเทมเพลตเป็นการ
    # จำลองชั่วคราวก่อนมีหน้าเว็บ ตอนนี้มีหน้าเว็บให้กรอกแล้ว (buffer_override) จะใช้ค่านั้นเป็นหลัก
    if buffer_override is not None:
        buffer_qty_by_barcode = buffer_override
        print(f"[excel_export] ใช้ยอดเผื่อที่ Admin กรอกผ่านเว็บ {len(buffer_qty_by_barcode)} SKU")
    else:
        from customers.cpall.logic.logistic_plan_export import read_buffer_qty_from_template
        try:
            buffer_qty_by_barcode = read_buffer_qty_from_template()
            print(f"[excel_export] อ่านยอดเผื่อจากเทมเพลต 'รอบเช้าต่างจังหวัด' ได้ {len(buffer_qty_by_barcode)} SKU")
        except Exception as e:
            print(f"[excel_export] WARNING: อ่านยอดเผื่อไม่สำเร็จ ({e}) — จะไม่กรอกยอดเผื่อให้ในรอบนี้")
            buffer_qty_by_barcode = {}

    for barcode, sub_qty in qty_by_barcode.items():
        row = header_rows.get(barcode)
        if row is None:
            missing_in_template.append(barcode)
            continue

        grand_total = 0.0
        remaining = dict(sub_qty)  # ไว้เช็คว่ามี sub_location ไหนที่หาคอลัมน์ในเทมเพลตไม่เจอบ้าง

        for col, sub_location in col_to_sub_location.items():
            if sub_location not in covered_sub_locations:
                continue  # จุดนี้ยังไม่มีรอบ PO มาเลย -> เว้นว่างไว้ตามเดิม ไม่ใช่ 0
            qty = sub_qty.get(sub_location, 0)
            # Production Plan (ต่างจาก Logistic Plan) ถ้ายอดสั่งเป็น 0 ให้เว้นว่างไว้ ไม่เขียน 0 ลงไป
            # (ตามที่ Admin ต้องการ — Logistic Plan/แผนรถยังคงเขียน 0 ชัดเจนเหมือนเดิม) — ต้องเคลียร์
            # เซลล์ผ่าน .value ตรงๆ เสมอ (ไม่ใช่แค่ "ข้ามไม่เขียน") เพราะไฟล์เทมเพลต live ถูกใช้ซ้ำทุก
            # รอบ ถ้าเคยมีค่าจากรอบก่อนหน้าค้างอยู่ (ไม่ใช่ 0) แล้วรอบนี้ไม่เขียนอะไรเลย ค่าเก่าจะยังคง
            # ค้างอยู่โดยไม่ตั้งใจ (เจอบั๊กจริงแบบนี้กับยอดเผื่อมาแล้ว — ดู comment ด้านล่างเรื่องยอดเผื่อ)
            ws.cell(row=row, column=col).value = qty if qty != 0 else None
            grand_total += qty
            remaining.pop(sub_location, None)

        # หมายเหตุ: ไม่เขียนทับคอลัมน์ "ยอดรวม" (total_col) — เทมเพลตเดิมมีสูตร =SUM(G:Q) อยู่แล้ว
        # ในแถวนั้น ถ้าเขียนค่านิ่งทับจะลบสูตรทิ้ง ทำให้ยอดไม่อัปเดตอัตโนมัติถ้า Admin แก้เลขทีหลัง
        # grand_total ที่คำนวณไว้ใช้แค่ log ตรวจสอบฝั่ง Python เท่านั้น

        unmatched_sub_locations.update(remaining.keys())
        filled.append(barcode)

        # เขียนยอดเผื่อเสมอ (ไม่ว่าจะมีข้อมูลสำหรับ SKU นี้หรือไม่) — ***บั๊กที่เคยเจอ***: เดิมเขียนแค่
        # ตอนมีค่า (if buffer_qty is not None) ทำให้ตอนไม่มีรอบเช้าต่างจังหวัด (buffer_override={})
        # ไม่เขียนอะไรเลย ปล่อยให้ "ค่ายอดเผื่อเก่าที่เคยเขียนไว้ในรอบก่อนหน้า" ยังค้างอยู่ในไฟล์เทมเพลต
        # live (ไฟล์เดียวกันถูกใช้ซ้ำทุกรอบ) ทำให้ยอดที่ต้องผลิตจริงคำนวณผิดจากยอดเผื่อที่ไม่เกี่ยวข้อง
        # กับรอบนี้เลย — ต้องเขียนเสมอ ถ้าไม่มีค่า (None) ก็เขียน None ทับ เพื่อล้างค่าเก่าทิ้งให้ชัดเจน
        #
        # ***บั๊กที่ 2 ที่เพิ่งเจอ*** (สำคัญกว่า): ws.cell(row=.., column=.., value=None) ของ openpyxl
        # "ไม่เคลียร์" เซลล์จริง! — openpyxl ตีความ value=None ว่า "ไม่ได้ส่ง value มา" (เหมือนไม่ระบุ
        # parameter นี้เลย) ไม่ใช่ "ตั้งค่าเป็นว่างเปล่า" ทดสอบยืนยันแล้วว่าเซลล์ที่มีค่าเก่าอยู่ก่อน จะ
        # ยังคงค่าเดิมอยู่แม้เรียกแบบนี้ก็ตาม — ต้องเข้าถึง .value ผ่าน attribute ตรงๆ ถึงจะเคลียร์ได้จริง
        buffer_qty = buffer_qty_by_barcode.get(barcode)
        ws.cell(row=row + BUFFER_ROW_OFFSET, column=BUFFER_COL).value = buffer_qty

    if missing_in_template:
        # ไม่ใช่แค่เตือน — หยุดทันที เพราะแปลว่ามี SKU สั่งจริงใน PO แต่จะหายไปเงียบๆ จากไฟล์ผลลัพธ์
        # (สาเหตุที่เจอบ่อย: มี SKU ใหม่ที่ยังไม่เคยมีในไฟล์เทมเพลตนี้มาก่อน) — ไม่ save ไฟล์ที่ไม่ครบออกไป
        msg_lines = [f"พบ {len(missing_in_template)} SKU ที่มีออเดอร์จริงใน PO แต่หาแถวใน Template ไม่เจอ:"]
        for b in missing_in_template:
            msg_lines.append(f"    - {b}")
        msg_lines.append(f"  -> ไปเพิ่มแถว SKU นี้ในไฟล์เทมเพลต {template_path} ก่อน (คัดลอกรูปแบบแถวอื่นที่มีอยู่) แล้วรันใหม่")
        raise ExcelExportError("\n".join(msg_lines))

    wb.save(output_path)

    print(f"[excel_export] เขียนไฟล์ {output_path} — กรอกยอดครบ {len(filled)}/{len(qty_by_barcode)} SKU")
    if unmatched_sub_locations:
        print(f"[excel_export] WARNING: พบจุดส่งย่อยที่มีข้อมูล แต่หาคอลัมน์ในเทมเพลตไม่เจอ "
              f"(เช็ค SUB_LOCATION_LABEL_CORRECTIONS): {unmatched_sub_locations}")

    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python -m src.excel_export <po_import_id> [<po_import_id_2> ...] <output_path>")
        sys.exit(1)

    *import_ids_str, output_path = sys.argv[1:]
    export_production_plan([int(x) for x in import_ids_str], output_path)
