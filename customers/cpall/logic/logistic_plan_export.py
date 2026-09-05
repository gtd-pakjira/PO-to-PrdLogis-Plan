"""
logistic_plan_export.py — Module 5: Logistic Plan Calculator + Exporter

ขอบเขตตอนนี้ (ตามที่ตกลงกัน): กรอกแค่ "ยอดสั่งตาม PO" ต่อ SKU x จุดส่งย่อย x PO ให้ถูกต้อง
ยังไม่ทำเรื่องจัดรถ/box-fill (ชีต "-รถ" ในแต่ละไฟล์ปล่อยไว้ตามเดิมทั้งหมด ไม่แตะเลย)

*** สำคัญมาก — ต้องอ่านก่อนใช้งาน ***
จำนวนคอลัมน์ PO ต่อจุดส่งย่อยในเทมเพลตแต่ละไฟล์ "ไม่คงที่" — ขึ้นกับว่าตอนสร้างเทมเพลตรอบนั้น
มีกี่ PO จริงสำหรับจุดส่งนั้น (เช่น รอบตัวอย่างนี้ สุวรรณภูมิได้ 10 PO เทมเพลตก็มี 10 คอลัมน์ PO1-PO10)
ระบบนี้ "ไม่ได้สร้างคอลัมน์ใหม่ให้เอง" — ใช้จำนวนคอลัมน์ที่มีอยู่ในเทมเพลตเป็นเพดาน:
  - ถ้า PO รอบนี้น้อยกว่าคอลัมน์ที่มี -> เว้นคอลัมน์ที่เหลือว่างไว้ (ไม่ error)
  - ถ้า PO รอบนี้มากกว่าคอลัมน์ที่มี -> "หยุดทันทีและแจ้งเตือน" ให้ Admin ไปเพิ่มคอลัมน์ในเทมเพลตเอง
    (ไม่เดาสร้างคอลัมน์ใหม่ให้ เพราะจะไปทำลาย format/สูตรของเทมเพลตเดิม)

*** สมมติฐานที่ยังไม่ได้ยืนยันกับ Admin — ต้องเอาไปถามทีหลัง ***
ไม่มีข้อมูลอะไรในไฟล์ PO ที่บอกว่า PO ใบไหนควรอยู่ในคอลัมน์ "PO1" ใบไหนอยู่ "PO2" (วันที่/เวลาส่ง
เหมือนกันหมดทุกใบในจุดส่งเดียวกัน) ระบบนี้จึง "เรียงตามเลขที่ PO จากน้อยไปมาก" แล้ว map เข้า
PO1, PO2, PO3... ตามลำดับนั้น เป็นการเดาที่มีเหตุผลรองรับ (เลข PO ดูเรียงตามลำดับที่ออกจริง)
แต่ยังไม่ได้ยืนยันว่า Admin ใช้กติกาเดียวกันหรือเปล่า — ต้องถามให้แน่ใจก่อนใช้งานจริง

วิธีรัน (จาก root ของโปรเจกต์):
    python -m src.logistic_plan_export <po_import_id> <group_name> <output_path>
    เช่น: python -m src.logistic_plan_export 1 บางบัวทอง cpall/data/output/บางบัวทอง.xlsx
"""
import re
from collections import defaultdict

import openpyxl

from customers.cpall.logic.date_utils import fixed_date_resolver, update_date_headers
from customers.cpall.logic.grouping import get_grouped_quantities_by_sub_location_and_po


def get_po_number_by_column_label(po_import_ids: list[int], group_name: str) -> dict:
    """
    คืน {"ขอนแก่น PO1": "F082833308", ...} — เลข PO จริงที่อยู่เบื้องหลังแต่ละ column_label
    ("PO1"/"PO2" ในตารางเป็นแค่ลำดับตำแหน่งในเทมเพลต ไม่ใช่เลข PO จริง) ใช้ logic การเรียงลำดับ
    เดียวกับตอน export_logistic_plan() เป๊ะ (เรียง po_number จากน้อยไปมากต่อ sub_location) เพื่อให้
    ได้ผลตรงกับที่ใช้ตอนสร้างแผนจริง — ใช้แสดง tooltip ในตารางเว็บเท่านั้น ไม่กระทบการคำนวณอะไรเลย
    """
    from customers.cpall.models import LocationMapping

    raw = get_grouped_quantities_by_sub_location_and_po(po_import_ids)
    # "group_name" อาจมีหลาย sub_location ในกลุ่มเดียว (เช่น "รอบเช้าต่างจังหวัด") — ดึงจาก
    # location_mapping ตรงๆ แทนที่จะเดาจากชื่อกลุ่ม
    group_sub_locations = set(
        LocationMapping.objects.filter(group=group_name).values_list("sub_location", flat=True)
    )

    po_numbers_by_sub_location = defaultdict(set)
    for row in raw:
        if row["sub_location"] in group_sub_locations:
            po_numbers_by_sub_location[row["sub_location"]].add(row["po_number"])

    result = {}
    for sub_loc, po_numbers in po_numbers_by_sub_location.items():
        sorted_pos = sorted(po_numbers)
        for i, po_number in enumerate(sorted_pos, start=1):
            result[f"{sub_loc} PO{i}"] = po_number
    return result


def get_group_templates() -> dict:
    """
    คืน dict รูปแบบเดียวกับ GROUP_TEMPLATES เดิม (เพื่อไม่ต้องแก้โค้ดที่ใช้อยู่เดิมเยอะเกินจำเป็น):
    {group_name: (template_path, sheet_name)} — แต่ query สดจาก database ทุกครั้งแทนที่จะ hardcode
    ไว้เป็น module constant ตอน import — เพิ่ม/แก้/ปิดใช้งานกลุ่มพื้นที่ผ่านหน้าเว็บได้เลย ไม่ต้องแก้
    โค้ด+deploy ใหม่แล้ว (เดิมต้องแก้ GROUP_TEMPLATES ในไฟล์นี้ตรงๆ ถ้าจะเพิ่มกลุ่มที่ 5)

    ไม่ cache ไว้ตรงๆ เพราะ query นี้เบามาก (ไม่กี่แถว) และการันตีว่าเห็นข้อมูลล่าสุดเสมอ (เช่น ถ้า
    Admin เพิ่งปิดใช้งานกลุ่มไปหมาดๆ) สำคัญกว่าการประหยัด query 1 ครั้ง
    """
    from customers.cpall.models import LogisticGroup

    result = {}
    for g in LogisticGroup.objects.filter(is_active=True):
        # ไฟล์ "live" ของเทมเพลตนี้ (path ตายตัวที่ export logic ทั้งหมดอ่านจากตรงนี้เสมอ) — ต้องตรง
        # กับ pattern เดิมที่ไฟล์จริงมีอยู่แล้ว: "logistic_plan_{ชื่อกลุ่ม}.xlsx" (ตรงกับที่
        # template_manager.py ใช้ sync ไฟล์ live จาก TemplateVersion ที่ active อยู่ด้วย)
        live_path = f"customers/cpall/excel_templates/logistic_plan_{g.group_name}.xlsx"
        result[g.group_name] = (live_path, g.sheet_name)
    return result


PO_LABEL_RE = re.compile(r"^PO\s*(\d+)$", re.IGNORECASE)
HEADER_SEARCH_ROWS_ABOVE = 6   # ค้นหาป้ายกำกับ (PO n / ชื่อจุดส่งย่อย / แพค) ในกี่แถวเหนือแถวหัว SKU
BARCODE_RE = re.compile(r"\d{10,16}")

# บาร์โค้ดที่เทมเพลตแต่ละไฟล์พิมพ์ผิด (พบระหว่างทดสอบ Module 5 — แต่ละไฟล์มีจุดผิดของตัวเอง แยกจาก
# ที่เจอตอนทำ Production Plan เพราะเป็นไฟล์คนละไฟล์กัน)
LEGACY_LOGISTIC_TEMPLATE_BARCODE_CORRECTIONS = {
    "885938000018": "8859388000018",   # องุ่นไชน์มัสแคท 300 กรัม (มหาชัย.xlsx — พิมพ์ขาด 1 หลัก)
    "88859388000063": "8859388000063",  # องุ่นเคียวโฮ 300 กรัม (สุวรรณภูมิ.xlsx — พิมพ์เกิน 1 หลัก)
    "88593880000189": "8859388000018",  # องุ่นไชน์มัสแคท 300 กรัม (รอบเช้าต่างจังหวัด.xlsx — มีเลข 9 เกินท้าย)
    "888593880000063": "8859388000063",  # องุ่นเคียวโฮ 300 กรัม (รอบเช้าต่างจังหวัด.xlsx — มีเลข 8 เกินหน้า)
}
# ชื่อ SKU ที่บาร์โค้ดในเทมเพลตเป็นเลขจริงของ SKU อื่น (สลับกัน ไม่ใช่แค่พิมพ์ผิด) — เช็คจากชื่อแทน
LEGACY_LOGISTIC_TEMPLATE_NAME_OVERRIDES = {
    "องุ่นดำไร้เมล็ด 300": "8859388002562",  # เทมเพลตใส่บาร์โค้ดของสาลี่ทองแพ็ก 2 ผล มาแทน (เหมือนไฟล์ Production Plan)
}

# เทมเพลตสะกดชื่อจุดส่งย่อยไม่ตรงกับ location_mapping.yaml (เจอระหว่างทำ Production Plan — ใช้ชุดเดียวกัน)
SUB_LOCATION_LABEL_CORRECTIONS = {
    "สุราษร์": "สุราษฎร์ธานี",
    "นครราขสีมา": "นครราชสีมา",
    "ชลุบรี": "ชลบุรี",  # เจอใหม่ในไฟล์บางบัวทอง.xlsx (พิมพ์ผิด)
    "ขอนแก่น  KK": "ขอนแก่น",  # ไฟล์รอบเช้าต่างจังหวัด — มีรหัสจังหวัดต่อท้ายชื่อ
    "สุราษร์ธานี   ST": "สุราษฎร์ธานี",
    "เชียงใหม่  CM": "เชียงใหม่",
    "ภูเก็ต  PK": "ภูเก็ต",
    "หาดใหญ่  HY": "หาดใหญ่",
}


class LogisticPlanError(Exception):
    pass


def _get_sub_locations_for_group(group_name: str) -> set:
    """ดึง sub_location ทั้งหมดที่เป็นสมาชิกของกลุ่มพื้นที่นี้ จาก location_mapping"""
    from customers.cpall.logic.db import get_connection as _get_connection
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                'SELECT sub_location FROM location_mapping WHERE "group" = %s',
                (group_name,),
            )
            return {r[0] for r in cur.fetchall()}
    finally:
        conn.close()


def group_has_data(po_import_ids, group_name: str) -> bool:
    """เช็คว่ากลุ่มพื้นที่นี้มีข้อมูลจริงใน po_import_id(s) ที่ระบุไหม (ใช้ตัดสินใจว่าควร export ไฟล์นี้หรือข้าม)"""
    if isinstance(po_import_ids, int):
        po_import_ids = [po_import_ids]
    group_subs = _get_sub_locations_for_group(group_name)
    raw = get_grouped_quantities_by_sub_location_and_po(po_import_ids)
    return any(r["sub_location"] in group_subs for r in raw)


def _find_line_no_column(ws, search_rows=range(1, 60), search_cols=range(1, 10)):
    """
    หาคอลัมน์ 'ลำดับ' (บรรทัดแรกของ SKU แต่ละตัว = เลข 1, 2, 3, ...)
    เช็คว่าอีก 2 คอลัมน์ถัดไปเป็นตัวเลข (บรรจุ/ตก.) เพื่อลดโอกาสจับผิดคอลัมน์
    """
    for col in search_cols:
        for row in search_rows:
            if ws.cell(row=row, column=col).value == 1:
                pack_cell = ws.cell(row=row, column=col + 2).value
                if isinstance(pack_cell, (int, float)) and 1 <= pack_cell <= 200:
                    return col, row
    raise LogisticPlanError("หาคอลัมน์ 'ลำดับ' (เลข 1 แถวแรกของ SKU) ในเทมเพลตไม่เจอ")


def _find_qty_column_range(ws, qty_start_col, header_row):
    """หาคอลัมน์สุดท้ายของยอดสั่ง โดยหาคำว่า 'แพค' ที่อยู่ถัดจากคอลัมน์ยอดสั่งคอลัมน์สุดท้าย"""
    for row in range(max(1, header_row - HEADER_SEARCH_ROWS_ABOVE), header_row + 2):
        for col in range(qty_start_col, qty_start_col + 30):
            val = ws.cell(row=row, column=col).value
            if val and "แพค" in str(val):
                return qty_start_col, col - 1
    raise LogisticPlanError("หาคอลัมน์ 'แพค' (ตัวบอกจุดจบของคอลัมน์ยอดสั่ง) ในเทมเพลตไม่เจอ")


def _find_column_labels(ws, col, header_row):
    """
    หา (sub_location, po_index) ของคอลัมน์หนึ่ง โดยไล่ดูแถวเหนือแถวหัว SKU ขึ้นไป
    - เจอ pattern 'PO n' -> po_index = n
    - เจอข้อความไทยอื่นที่ไม่ใช่เวลา/PO/หัวเรื่องวันที่ -> ใช้เป็น sub_location
    คืนค่า sub_location = None ถ้าหาไม่เจอในคอลัมน์นี้ (ปกติเกิดกับคอลัมน์ PO2, PO3, ... ที่ตามหลัง
    คอลัมน์แรกของจุดส่งย่อยเดียวกัน เพราะเทมเพลตใช้ merged cell ใส่ชื่อแค่ครั้งเดียว) — ผู้เรียกต้อง
    forward-fill จากคอลัมน์ก่อนหน้าเอง
    """
    sub_location = None
    po_index = None
    for row in range(header_row - 1, max(0, header_row - HEADER_SEARCH_ROWS_ABOVE), -1):
        val = ws.cell(row=row, column=col).value
        if not val:
            continue
        text = str(val).strip()
        m = PO_LABEL_RE.match(text)
        if m and po_index is None:
            po_index = int(m.group(1))
            continue
        if sub_location is None and not re.match(r"^\d{1,2}[.:]\d{2}", text) and "ตี" not in text \
                and "น." not in text and "วันที่" not in text and "ส่ง" not in text:
            sub_location = SUB_LOCATION_LABEL_CORRECTIONS.get(text, text)

    return sub_location, po_index


def _find_sku_header_rows(ws, name_col):
    """หา (barcode -> แถวหัว SKU) — บาร์โค้ดอยู่แถวถัดไป (offset +1) ในคอลัมน์ชื่อ SKU"""
    line_no_col = name_col - 1
    mapping = {}
    for row in range(1, ws.max_row + 1):
        line_no = ws.cell(row=row, column=line_no_col).value
        if not isinstance(line_no, (int, float)):
            continue
        header_name = str(ws.cell(row=row, column=name_col).value or "")
        barcode_cell = ws.cell(row=row + 1, column=name_col).value
        if not barcode_cell:
            continue

        matched_by_name = False
        for name_fragment, correct_barcode in LEGACY_LOGISTIC_TEMPLATE_NAME_OVERRIDES.items():
            if name_fragment in header_name:
                mapping[correct_barcode] = row
                matched_by_name = True
                break
        if matched_by_name:
            continue

        m = BARCODE_RE.search(str(barcode_cell))
        if m:
            raw_barcode = m.group(0)
            corrected = LEGACY_LOGISTIC_TEMPLATE_BARCODE_CORRECTIONS.get(raw_barcode, raw_barcode)
            mapping[corrected] = row
    return mapping


def _find_buffer_column(ws) -> int:
    """
    หาคอลัมน์ 'ยอดเผื่อ' จากหัวตาราง (row12 ของเทมเพลตนี้มีคำว่า 'ยอดเผื่อ' ชัดเจน)
    ใช้เฉพาะไฟล์เทมเพลตที่มีคอลัมน์นี้จริง (ตอนนี้มีแค่ไฟล์ รอบเช้าต่างจังหวัด ไฟล์เดียว —
    อีก 3 ไฟล์ไม่มีคอลัมน์ยอดเผื่อเลย)
    """
    for row in range(1, 15):
        for col in range(1, 30):
            val = ws.cell(row=row, column=col).value
            if val and "ยอดเผื่อ" in str(val):
                return col
    raise LogisticPlanError("หาคอลัมน์ 'ยอดเผื่อ' ในเทมเพลตนี้ไม่เจอ")


def read_buffer_qty_from_template(group_name: str = "รอบเช้าต่างจังหวัด") -> dict:
    """
    อ่านค่ายอดเผื่อที่ Admin กรอกไว้ในไฟล์เทมเพลตของกลุ่มนี้ (Admin แก้ไฟล์เทมเพลตตรงๆ ตอนนี้ —
    ยังไม่มีหน้าเว็บให้กรอก รอทำตอน Flask) คืนค่า {barcode: buffer_qty}

    *** หมายเหตุสำคัญ ***
    ตอนนี้มีแค่กลุ่ม "รอบเช้าต่างจังหวัด" เท่านั้นที่มีคอลัมน์ยอดเผื่อในเทมเพลต (อีก 3 กลุ่มไม่มี)
    และยังไม่รู้ว่า Admin คำนวณยอดเผื่อยังไง (คงที่ต่อ SKU ไม่ขึ้นกับกลุ่มพื้นที่ หรือมีเงื่อนไขอื่น)
    — ตอนนี้แค่จำลองว่า "เอาค่าที่กรอกไว้ในไฟล์เทมเพลตนี้ไปใช้กับทั้ง Production Plan" ตามที่ตกลงกัน
    ต้องยืนยันกับ Admin จริงอีกทีว่าใช้หลักการนี้ถูกไหม
    """
    group_templates = get_group_templates()
    if group_name not in group_templates:
        raise LogisticPlanError(f"ไม่รู้จักกลุ่มพื้นที่ '{group_name}'")

    template_path, sheet_name = group_templates[group_name]
    wb = openpyxl.load_workbook(template_path, data_only=True)  # data_only เพราะยอดเผื่อเป็นค่าคงที่ ไม่ใช่สูตร
    ws = wb[sheet_name]

    buffer_col = _find_buffer_column(ws)
    line_no_col, header_row = _find_line_no_column(ws)
    name_col = line_no_col + 1
    header_rows = _find_sku_header_rows(ws, name_col)

    result = {}
    for barcode, row in header_rows.items():
        val = ws.cell(row=row, column=buffer_col).value
        if val:
            result[barcode] = float(val)

    return result


def export_logistic_plan(po_import_ids, group_name: str, output_path: str):
    """
    po_import_ids: รับได้ทั้ง int เดี่ยว หรือ list ของ int
    วันที่ในหัวไฟล์: ดึงจากวันที่ที่ผูกไว้กับรอบ PO ที่มีข้อมูลของกลุ่มนี้ (ไฟล์นี้มาจากรอบเดียวเสมอ
    ในทางปฏิบัติ เพราะจุดส่งย่อยของกลุ่มหนึ่งอยู่ในรอบ PO เดียวกันหมด)
    """
    if isinstance(po_import_ids, int):
        po_import_ids = [po_import_ids]

    group_templates = get_group_templates()
    if group_name not in group_templates:
        raise LogisticPlanError(f"ไม่รู้จักกลุ่มพื้นที่ '{group_name}' (ต้องเป็นหนึ่งใน {list(group_templates)})")

    template_path, sheet_name = group_templates[group_name]
    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name]

    from customers.cpall.logic.grouping import get_dates_by_sub_location
    group_sub_locations_for_dates = _get_sub_locations_for_group(group_name)
    dates_by_sub_location = get_dates_by_sub_location(po_import_ids)
    # เอาวันที่ของจุดส่งย่อยจุดแรกในกลุ่มนี้ที่มีข้อมูล (ทุกจุดในกลุ่มเดียวกันควรมาจากรอบเดียวกันอยู่แล้ว)
    plan_dates = next(
        (dates_by_sub_location[s] for s in group_sub_locations_for_dates if s in dates_by_sub_location),
        (None, None),
    )
    if plan_dates[0] and plan_dates[1]:
        n = update_date_headers(ws, fixed_date_resolver(*plan_dates))
        print(f"[logistic_plan_export:{group_name}] อัปเดตวันที่ในหัวไฟล์ {n} จุด "
              f"(ผลิต={plan_dates[0]}, PO={plan_dates[1]})")

    line_no_col, header_row = _find_line_no_column(ws)
    name_col = line_no_col + 1
    qty_start_col = line_no_col + 3
    qty_start_col, qty_end_col = _find_qty_column_range(ws, qty_start_col, header_row)

    # map คอลัมน์ -> (sub_location, po_index)
    # คอลัมน์ PO2/PO3/... ที่ตามหลังคอลัมน์แรกของจุดส่งย่อยเดียวกันมักไม่มีป้ายชื่อซ้ำ (merged cell)
    # -> forward-fill จากคอลัมน์ก่อนหน้า ถ้าคอลัมน์นี้หา sub_location เองไม่เจอ
    col_labels = {}
    last_sub_location = None
    for col in range(qty_start_col, qty_end_col + 1):
        sub_loc, po_idx = _find_column_labels(ws, col, header_row)
        if sub_loc is None:
            sub_loc = last_sub_location if last_sub_location is not None else group_name
        last_sub_location = sub_loc
        col_labels[col] = (sub_loc, po_idx)

    # ถ้ามีแค่คอลัมน์เดียวและไม่เจอ po_index -> ตั้งเป็น 1 (เคสจุดส่งเดียว 1 PO เช่น มหาชัย)
    if len(col_labels) == 1:
        only_col = list(col_labels.keys())[0]
        sub_loc, po_idx = col_labels[only_col]
        if po_idx is None:
            col_labels[only_col] = (sub_loc, 1)

    print(f"[logistic_plan_export:{group_name}] พบคอลัมน์ยอดสั่ง {qty_start_col}-{qty_end_col} "
          f"({qty_end_col - qty_start_col + 1} คอลัมน์)")
    for col, (sub_loc, po_idx) in col_labels.items():
        print(f"    col {col}: sub_location={sub_loc}, PO{po_idx}")

    # จัดกลุ่มคอลัมน์ตาม sub_location เพื่อจับคู่กับจำนวน PO จริงทีหลัง
    cols_by_sub_location = defaultdict(list)
    for col, (sub_loc, po_idx) in col_labels.items():
        cols_by_sub_location[sub_loc].append((po_idx, col))
    for sub_loc in cols_by_sub_location:
        cols_by_sub_location[sub_loc].sort(key=lambda x: (x[0] is None, x[0]))

    # ---------- ดึงข้อมูลจริงจาก DB (กรองเฉพาะจุดส่งย่อยที่เป็นของกลุ่มนี้เท่านั้น) ----------
    group_sub_locations = _get_sub_locations_for_group(group_name)
    raw = get_grouped_quantities_by_sub_location_and_po(po_import_ids)
    raw = [r for r in raw if r["sub_location"] in group_sub_locations]

    by_sub_location = defaultdict(lambda: defaultdict(dict))  # sub_location -> barcode -> {po_number: qty}
    po_numbers_by_sub_location = defaultdict(set)
    for row in raw:
        by_sub_location[row["sub_location"]][row["barcode"]][row["po_number"]] = row["qty_case_ordered"]
        po_numbers_by_sub_location[row["sub_location"]].add(row["po_number"])

    # ---------- เช็คว่าคอลัมน์ในเทมเพลตพอไหม ----------
    overflow = []
    for sub_loc, po_numbers in po_numbers_by_sub_location.items():
        available_cols = len(cols_by_sub_location.get(sub_loc, []))
        if len(po_numbers) > available_cols:
            overflow.append((sub_loc, len(po_numbers), available_cols))
    if overflow:
        msg_lines = [f"เทมเพลต '{template_path}' มีคอลัมน์ PO ไม่พอสำหรับรอบนี้:"]
        for sub_loc, needed, available in overflow:
            msg_lines.append(f"    - {sub_loc}: มี {needed} PO จริง แต่เทมเพลตรองรับได้แค่ {available} คอลัมน์")
        msg_lines.append("  -> ไปเพิ่มคอลัมน์ PO ในไฟล์เทมเพลตนี้ก่อน (คัดลอกรูปแบบคอลัมน์ PO ล่าสุด) แล้วรันใหม่")
        raise LogisticPlanError("\n".join(msg_lines))

    # ---------- เรียง PO ตามเลขที่ PO จากน้อยไปมาก แล้วจับคู่เข้าคอลัมน์ PO1, PO2, ... ----------
    # (ดูคำเตือนบนสุดของไฟล์นี้ — สมมติฐานที่ยังไม่ได้ยืนยันกับ Admin)
    # เรียงตาม "sub_location" (ไม่ใช่ต่อ SKU) เพื่อให้ตำแหน่ง PO1/PO2/PO3 คงที่ทุก SKU ในจุดเดียวกัน
    # (ถ้าเรียงต่อ SKU เอง จะเลื่อนตำแหน่งผิดเวลา SKU นั้นไม่ได้สั่งใน PO ตัวใดตัวหนึ่ง)
    global_sorted_po_by_sub_location = {
        sub_loc: sorted(po_numbers) for sub_loc, po_numbers in po_numbers_by_sub_location.items()
    }

    header_rows = _find_sku_header_rows(ws, name_col)
    filled_skus, missing_in_template = set(), set()

    # สำคัญ: เคลียร์ทุกช่องยอดสั่ง (qty_start_col..qty_end_col) ของทุกแถว SKU ในเทมเพลตนี้ก่อนเขียนใหม่
    # เสมอ — บางไฟล์เทมเพลตไม่ใช่ไฟล์เปล่า แต่มีตัวเลขจากรอบก่อนหน้ากรอกค้างอยู่จริง ถ้าไม่เคลียร์ก่อน
    # ช่องที่ยอดใหม่รอบนี้เป็น 0 (ไม่มีออเดอร์) จะยังโชว์ตัวเลขเก่าค้างอยู่ ทำให้ยอดรวมผิด
    for row in header_rows.values():
        for col in range(qty_start_col, qty_end_col + 1):
            ws.cell(row=row, column=col).value = None

    # เขียนทุก SKU ที่เจอในเทมเพลต (ไม่ใช่แค่ SKU ที่มีออเดอร์) — จุดส่งย่อยไหนที่ "มีรอบ PO มาแล้ว"
    # (อยู่ใน po_numbers_by_sub_location) ให้กรอก 0 ชัดเจนถ้า SKU นั้นสั่ง 0 ที่จุดนั้น
    for barcode, row in header_rows.items():
        for sub_loc, sorted_pos in global_sorted_po_by_sub_location.items():
            available_cols = cols_by_sub_location.get(sub_loc)
            if not available_cols:
                continue
            po_qty_map = by_sub_location.get(sub_loc, {}).get(barcode, {})
            for po_number, col_pair in zip(sorted_pos, available_cols):
                po_idx, col = col_pair
                qty = po_qty_map.get(po_number, 0)
                ws.cell(row=row, column=col).value = qty
        filled_skus.add(barcode)

    # เช็คว่ามี barcode ที่มีออเดอร์จริง แต่หาแถวในเทมเพลตไม่เจอบ้างไหม (SKU ใหม่ที่ยังไม่มีใน template)
    all_ordered_barcodes = {r["barcode"] for r in raw}
    for barcode in all_ordered_barcodes:
        if barcode not in header_rows:
            missing_in_template.add(barcode)

    if missing_in_template:
        # ไม่ใช่แค่เตือน — หยุดทันที เพราะแปลว่ามี SKU สั่งจริงใน PO แต่จะหายไปเงียบๆ จากไฟล์ผลลัพธ์
        # (เช่น รอบนี้จู่ๆ มีคนสั่ง SKU ที่ไฟล์เทมเพลตกลุ่มนี้ไม่เคยมีแถวไว้มาก่อน — ไฟล์รอบเช้าต่างจังหวัด
        # ตอนนี้มีแค่ 18 SKU ไม่ครบ 19 เหมือนกลุ่มอื่น จุดนี้จะโดนจับได้ตรงนี้พอดี) — ไม่ save ไฟล์ที่ไม่ครบออกไป
        msg_lines = [f"พบ {len(missing_in_template)} SKU ที่มีออเดอร์จริงใน PO แต่หาแถวในเทมเพลต '{template_path}' ไม่เจอ:"]
        for b in missing_in_template:
            msg_lines.append(f"    - {b}")
        msg_lines.append("  -> ไปเพิ่มแถว SKU นี้ในไฟล์เทมเพลต (คัดลอกรูปแบบแถวอื่นที่มีอยู่) แล้วรันใหม่")
        raise LogisticPlanError("\n".join(msg_lines))

    # สินค้าที่ปิดใช้งาน (is_active=False) และไม่มี PO สั่งเลยในกลุ่มนี้รอบนี้ แต่ยังมีแถวอยู่ในเทมเพลต —
    # ซ่อนแถวไว้ (ไม่ลบจริง) กันสูตรที่ reference row number อื่นในเทมเพลตพัง — ถ้า inactive แต่ยังมี
    # PO สั่งอยู่จริง จะไม่มาถึงจุดนี้เลย เพราะ run_plan() block ไปตั้งแต่ก่อนเรียกฟังก์ชันนี้แล้ว
    from customers.cpall.models import ProductMaster
    inactive_barcodes = set(
        ProductMaster.objects.filter(is_active=False).values_list("barcode", flat=True)
    )
    hidden_count = 0
    for barcode, row in header_rows.items():
        if barcode in inactive_barcodes and barcode not in all_ordered_barcodes:
            ws.row_dimensions[row].hidden = True
            ws.row_dimensions[row + 1].hidden = True  # แถวรอง (ชื่ออังกฤษ/pack breakdown)
            hidden_count += 1
    if hidden_count:
        print(f"[logistic_plan_export:{group_name}] ซ่อน {hidden_count} แถวสินค้าที่ปิดใช้งานและไม่มี PO สั่งในรอบนี้")

    wb.save(output_path)

    print(f"[logistic_plan_export:{group_name}] เขียนไฟล์ {output_path} — กรอกยอดครบ {len(filled_skus)} SKU")

    return output_path
