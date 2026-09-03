"""
template_manager.py — ดาวน์โหลด/อัปโหลด/ตรวจสอบไฟล์ Template (Production Plan + Logistic Plan 4 กลุ่ม)
พร้อมระบบ versioning เต็มรูปแบบ (Phase 1.6 sub-phase 1)

Admin แก้ไฟล์ Template เองบ่อย (เพิ่มคอลัมน์ PO, เพิ่มแถว SKU ใหม่, แก้สูตร) — โมดูลนี้:
  1. ตรวจสอบโครงสร้างไฟล์ใหม่ทันทีตอนอัปโหลด โดยใช้ตัวสแกนเดียวกับที่ excel_export.py /
     logistic_plan_export.py ใช้ตอนสร้างแผนจริง — ถ้าไฟล์ใหม่ผ่านตรงนี้ แปลว่าใช้สร้างแผนได้แน่นอน
  2. เก็บทุกเวอร์ชันที่เคยอัปโหลดไว้ถาวร (ไม่ใช่แค่ backup 1 ชั้นแบบเดิม) — กู้คืนไปเวอร์ชันไหนก็ได้
     ในประวัติ ลบเวอร์ชันเก่าได้ (ถ้าไม่ใช่เวอร์ชันสุดท้ายที่เหลือ และไม่มีแผนไหนอ้างอิงอยู่)

*** สถาปัตยกรรม: ยังมี "ไฟล์ live" ที่ path ตายตัวเดิม (get_template_registry()[key]["path"]) ***
excel_export.py / logistic_plan_export.py อ่านจาก path ตายตัวนี้เสมอ (ไม่ได้แก้ให้รู้จัก versioning
โดยตรง) — ทุกครั้งที่เปลี่ยนเวอร์ชัน active (อัปโหลดใหม่/กู้คืน) จะ sync ไฟล์ที่ path ตายตัวนี้ให้ตรงกับ
เวอร์ชัน active เสมอ วิธีนี้ทำให้โค้ดเดิมที่มีอยู่แล้วทำงานถูกต้องต่อไปโดยไม่ต้องแก้เลย
"""
import os
import shutil

import openpyxl
from openpyxl.utils import get_column_letter

from customers.cpall.logic.db import get_cpall_customer_id
from customers.cpall.logic.excel_export import SHEET_NAME as PP_SHEET_NAME
from customers.cpall.logic.excel_export import TEMPLATE_PATH as PP_TEMPLATE_PATH
from customers.cpall.logic.excel_export import _find_sku_header_rows as _find_pp_sku_header_rows
from customers.cpall.logic.excel_export import _find_sub_location_columns, _find_total_column
from customers.cpall.logic.logistic_plan_export import (
    _find_line_no_column,
    _find_qty_column_range,
    get_group_templates,
)
from customers.cpall.logic.logistic_plan_export import _find_sku_header_rows as _find_lp_sku_header_rows

VERSIONS_DIR = "customers/cpall/excel_templates/versions"


class TemplateValidationError(Exception):
    pass


class TemplateInUseError(Exception):
    """ลบเวอร์ชันนี้ไม่ได้ เพราะเป็นเวอร์ชันสุดท้ายที่เหลือ หรือมีแผนอ้างอิงอยู่"""
    pass


def get_template_registry() -> dict:
    """
    สร้าง registry ของ Template ทั้งหมด (Production Plan + Logistic Plan ทุกกลุ่ม) — query กลุ่มพื้นที่
    สดจาก database ทุกครั้ง (ผ่าน get_group_templates()) แทนที่จะ build ครั้งเดียวตอน import module
    เหมือนเดิม (TEMPLATE_REGISTRY เคยเป็น module-level constant) — เพิ่ม/ปิดใช้งานกลุ่มพื้นที่ผ่าน
    หน้าเว็บแล้วเห็นผลทันที ไม่ต้อง restart เซิร์ฟเวอร์
    """
    registry = {
        "production_plan": {
            "path": PP_TEMPLATE_PATH,
            "label": "Production Plan (แพลน 7-11)",
            "kind": "production",
        }
    }
    for group_name, (path, _) in get_group_templates().items():
        registry[f"logistic_{group_name}"] = {
            "path": path,
            "label": f"Logistic Plan — {group_name}",
            "kind": "logistic",
            "group": group_name,
        }
    return registry


def get_template_grid(key: str, sheet_name: str = None, max_rows: int = 120, max_cols: int = 25) -> dict:
    """
    อ่านไฟล์ Template แบบดิบๆ (ไม่คำนวณ ไม่แปลงอะไรเลย) มาจัดเป็นตารางสำหรับแสดงในเว็บ — เห็นสูตร
    Excel ตรงๆ (เช่น "=SUM(G11:Q11)") แทนที่จะเห็นแค่ค่าที่คำนวณแล้ว — อ่านจากไฟล์ live เสมอ (เวอร์ชัน
    active ปัจจุบัน) ไม่มีการแก้ไขไฟล์ใดๆ ในฟังก์ชันนี้ อ่านอย่างเดียว (read-only)

    sheet_name: ถ้าไม่ระบุ ใช้ชีตหลักของ template นี้ (ตามที่ลงทะเบียนไว้) — ไฟล์เดียวอาจมีหลายชีต
    (เช่น ชีต "-รถ"/"คันที่ 1/2" ที่เกี่ยวกับจัดรถ) เลือกดูชีตอื่นได้ผ่าน sheet_name
    """
    registry = get_template_registry()
    if key not in registry:
        raise TemplateValidationError(f"ไม่รู้จัก template '{key}'")

    info = registry[key]
    path = info["path"]
    if not os.path.exists(path):
        raise TemplateValidationError(f"ไม่พบไฟล์ {path}")

    wb = openpyxl.load_workbook(path, data_only=False)  # data_only=False สำคัญมาก — เอาไว้เห็นสูตรดิบ

    if info["kind"] == "production":
        default_sheet = PP_SHEET_NAME
    else:
        _, default_sheet = get_group_templates()[info["group"]]

    if sheet_name is None or sheet_name not in wb.sheetnames:
        sheet_name = default_sheet if default_sheet in wb.sheetnames else wb.sheetnames[0]

    ws = wb[sheet_name]
    n_rows = min(ws.max_row or 1, max_rows)
    n_cols = min(ws.max_column or 1, max_cols)

    col_letters = [get_column_letter(c) for c in range(1, n_cols + 1)]
    grid_rows = []
    for r in range(1, n_rows + 1):
        cells = []
        for c in range(1, n_cols + 1):
            val = ws.cell(row=r, column=c).value
            is_formula = isinstance(val, str) and val.startswith("=")
            cells.append({"value": val, "is_formula": is_formula})
        grid_rows.append({"row_num": r, "cells": cells})

    return {
        "label": info["label"],
        "sheet_names": wb.sheetnames,
        "current_sheet": sheet_name,
        "col_letters": col_letters,
        "rows": grid_rows,
        "truncated": (ws.max_row or 1) > max_rows or (ws.max_column or 1) > max_cols,
    }


def validate_template(key: str, filepath: str) -> dict:
    """
    ตรวจสอบโครงสร้างไฟล์ Template ที่อัปโหลดมา
    raise TemplateValidationError พร้อมเหตุผลชัดเจนถ้าไม่ผ่าน คืนค่าสรุปข้อมูลถ้าผ่าน (ไว้โชว์ผู้ใช้)
    """
    registry = get_template_registry()
    if key not in registry:
        raise TemplateValidationError(f"ไม่รู้จัก template '{key}'")

    info = registry[key]
    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        raise TemplateValidationError(f"เปิดไฟล์ไม่ได้ (ไม่ใช่ไฟล์ .xlsx ที่ถูกต้อง): {e}")

    if info["kind"] == "production":
        if PP_SHEET_NAME not in wb.sheetnames:
            raise TemplateValidationError(
                f"ไม่พบชีตชื่อ '{PP_SHEET_NAME}' ในไฟล์ — เช็คว่าไม่ได้เปลี่ยนชื่อชีตตอนแก้ไฟล์"
            )
        ws = wb[PP_SHEET_NAME]
        try:
            col_to_sub = _find_sub_location_columns(ws)
            total_col = _find_total_column(ws)
            header_rows = _find_pp_sku_header_rows(ws)
        except Exception as e:
            raise TemplateValidationError(f"อ่านโครงสร้างไฟล์ไม่ผ่าน: {e}")

        if not col_to_sub:
            raise TemplateValidationError("หาคอลัมน์จุดส่งย่อยไม่เจอ (แถวหัวตาราง 7-8 อาจถูกแก้ผิดรูปแบบ)")
        if total_col is None:
            raise TemplateValidationError("หาคอลัมน์ 'ยอดรวม' ไม่เจอ")
        if not header_rows:
            raise TemplateValidationError("หาแถว SKU ไม่เจอเลยสักแถว")

        return {"sub_locations": list(col_to_sub.values()), "sku_count": len(header_rows)}

    else:  # logistic
        group_name = info["group"]
        _, sheet_name = get_group_templates()[group_name]
        if sheet_name not in wb.sheetnames:
            raise TemplateValidationError(
                f"ไม่พบชีตชื่อ '{sheet_name}' ในไฟล์ — เช็คว่าไม่ได้เปลี่ยนชื่อชีตตอนแก้ไฟล์"
            )
        ws = wb[sheet_name]
        try:
            line_no_col, header_row = _find_line_no_column(ws)
            name_col = line_no_col + 1
            qty_start_col = line_no_col + 3
            qty_start_col, qty_end_col = _find_qty_column_range(ws, qty_start_col, header_row)
            header_rows = _find_lp_sku_header_rows(ws, name_col)
        except Exception as e:
            raise TemplateValidationError(f"อ่านโครงสร้างไฟล์ไม่ผ่าน: {e}")

        if not header_rows:
            raise TemplateValidationError("หาแถว SKU ไม่เจอเลยสักแถว")

        return {"qty_column_count": qty_end_col - qty_start_col + 1, "sku_count": len(header_rows)}


def _version_file_path(key: str, version_number: int) -> str:
    return os.path.join(VERSIONS_DIR, key, f"v{version_number}.xlsx")


def _sync_live_file(key: str, version):
    """คัดลอกไฟล์ของเวอร์ชันที่ระบุไปทับไฟล์ live ที่ path ตายตัว — ให้โค้ดเดิมอ่านถูกเวอร์ชันเสมอ"""
    shutil.copy2(version.file_path, get_template_registry()[key]["path"])


def _ensure_initial_version(key: str):
    """
    ถ้ายังไม่เคยมี TemplateVersion เลยสำหรับ key นี้ (ระบบเพิ่งอัปเกรดมาใช้ versioning ครั้งแรก)
    ให้สร้างเวอร์ชัน 1 จากไฟล์ live ปัจจุบันให้อัตโนมัติ — ไม่ต้องให้ Admin ทำอะไรเพิ่ม
    """
    from customers.cpall.models import TemplateVersion

    if TemplateVersion.objects.filter(template_key=key).exists():
        return

    target_path = get_template_registry()[key]["path"]
    if not os.path.exists(target_path):
        return  # ไม่มีไฟล์ live เลย ไม่มีอะไรให้สร้างเป็นเวอร์ชันแรก

    version_path = _version_file_path(key, 1)
    os.makedirs(os.path.dirname(version_path), exist_ok=True)
    shutil.copy2(target_path, version_path)

    try:
        validation = validate_template(key, target_path)
        summary = f"sku_count={validation.get('sku_count')}"
    except TemplateValidationError:
        summary = "(ไฟล์เดิมก่อนมีระบบ versioning — ไม่ได้ validate ตอนสร้างเวอร์ชันนี้)"

    TemplateVersion.objects.create(
        customer_id=get_cpall_customer_id(), template_key=key, version_number=1,
        file_path=version_path, is_active=True, validation_summary=summary,
    )


def list_templates() -> list[dict]:
    """คืนรายการ Template ทั้งหมด พร้อมข้อมูลเวอร์ชันปัจจุบัน — ใช้แสดงหน้าเว็บ"""
    from customers.cpall.models import TemplateVersion

    result = []
    for key, info in get_template_registry().items():
        _ensure_initial_version(key)
        active = TemplateVersion.objects.filter(template_key=key, is_active=True).first()
        total_versions = TemplateVersion.objects.filter(template_key=key).count()
        result.append({
            "key": key,
            "label": info["label"],
            "exists": active is not None,
            "active_version": active.version_number if active else None,
            "uploaded_at": active.uploaded_at if active else None,
            "total_versions": total_versions,
        })
    return result


def list_versions(key: str) -> list[dict]:
    """คืนทุกเวอร์ชันของ template นี้ เรียงใหม่สุดก่อน — ใช้แสดงหน้าประวัติเวอร์ชัน"""
    from customers.cpall.models import TemplateVersion

    _ensure_initial_version(key)
    versions = TemplateVersion.objects.filter(template_key=key).order_by("-version_number")
    return [
        {"id": v.id, "version_number": v.version_number, "is_active": v.is_active,
         "uploaded_at": v.uploaded_at, "validation_summary": v.validation_summary,
         "original_filename": v.original_filename}
        for v in versions
    ]


def upload_new_version(key: str, new_filepath: str, original_filename: str = None) -> dict:
    """
    ตรวจสอบไฟล์ใหม่ก่อน (validate_template) แล้วสร้างเป็นเวอร์ชันใหม่ + ตั้งเป็น active ทันที
    เวอร์ชันเก่าไม่ได้ถูกลบเลย แค่ไม่ active แล้ว (กู้คืนได้เสมอผ่าน restore_to_version)
    ถ้า validate ไม่ผ่าน จะ raise ทันที ไม่แตะเวอร์ชัน/ไฟล์ live เดิมเลย
    original_filename: ชื่อไฟล์ตอน Admin เลือกอัปโหลดจริง (เก็บไว้ให้ดูย้อนหลังในหน้าประวัติเวอร์ชัน)
    """
    if key not in get_template_registry():
        raise TemplateValidationError(f"ไม่รู้จัก template '{key}'")

    validation_result = validate_template(key, new_filepath)

    from customers.cpall.models import TemplateVersion

    _ensure_initial_version(key)
    last = TemplateVersion.objects.filter(template_key=key).order_by("-version_number").first()
    next_version = (last.version_number + 1) if last else 1

    version_path = _version_file_path(key, next_version)
    os.makedirs(os.path.dirname(version_path), exist_ok=True)
    shutil.move(new_filepath, version_path)

    TemplateVersion.objects.filter(template_key=key, is_active=True).update(is_active=False)
    new_version = TemplateVersion.objects.create(
        customer_id=get_cpall_customer_id(), template_key=key, version_number=next_version,
        file_path=version_path, original_filename=original_filename, is_active=True,
        validation_summary=f"sku_count={validation_result.get('sku_count')}",
    )
    _sync_live_file(key, new_version)

    return validation_result


def restore_to_version(key: str, version_id: int):
    """ตั้งเวอร์ชันที่ระบุให้กลับมาเป็น active (ไฟล์ของทุกเวอร์ชันยังอยู่ครบ แค่สลับว่าตัวไหน active)"""
    from customers.cpall.models import TemplateVersion

    try:
        version = TemplateVersion.objects.get(id=version_id, template_key=key)
    except TemplateVersion.DoesNotExist:
        raise TemplateValidationError("ไม่พบเวอร์ชันนี้")

    TemplateVersion.objects.filter(template_key=key, is_active=True).update(is_active=False)
    version.is_active = True
    version.save(update_fields=["is_active"])
    _sync_live_file(key, version)
    return version


def delete_version(key: str, version_id: int):
    """
    ลบเวอร์ชันที่ระบุถาวร (ลบทั้ง DB record และไฟล์บนดิสก์) — ลบไม่ได้ถ้า:
      1. เป็นเวอร์ชันสุดท้ายที่เหลืออยู่ของ template นี้ (ต้องมีอย่างน้อย 1 เวอร์ชันเสมอ)
      2. มีแผนที่เคยสร้างไว้ใช้เวอร์ชันนี้อยู่ (เช็คผ่าน plan_run.production_template_version /
         plan_run_logistic_file.template_version — คอลัมน์นี้เพิ่งเตรียมไว้ ยังไม่มีแผนไหนผูกจริง
         จนกว่าจะถึง sub-phase 3 ที่เปลี่ยน flow สร้างแผน ตอนนี้เช็คแล้วจะผ่านเสมอ แต่โค้ดพร้อมใช้)
    ถ้าลบเวอร์ชันที่ active อยู่ จะเลื่อนเวอร์ชันล่าสุดที่เหลือขึ้นมาเป็น active แทนอัตโนมัติ
    """
    from customers.cpall.models import PlanRun, PlanRunLogisticFile, TemplateVersion

    try:
        version = TemplateVersion.objects.get(id=version_id, template_key=key)
    except TemplateVersion.DoesNotExist:
        raise TemplateValidationError("ไม่พบเวอร์ชันนี้")

    total_count = TemplateVersion.objects.filter(template_key=key).count()
    if total_count <= 1:
        raise TemplateInUseError("ลบไม่ได้ — เหลือเทมเพลตนี้แค่เวอร์ชันเดียว ต้องมีอย่างน้อย 1 เวอร์ชันเสมอ")

    used_in_production = PlanRun.objects.filter(production_template_version=version).exists()
    used_in_logistic = PlanRunLogisticFile.objects.filter(template_version=version).exists()
    if used_in_production or used_in_logistic:
        raise TemplateInUseError("ลบไม่ได้ — มีแผนที่เคยสร้างไว้ใช้เทมเพลตเวอร์ชันนี้อยู่")

    was_active = version.is_active
    file_path = version.file_path
    version.delete()

    if os.path.exists(file_path):
        os.remove(file_path)

    if was_active:
        new_active = TemplateVersion.objects.filter(template_key=key).order_by("-version_number").first()
        if new_active:
            new_active.is_active = True
            new_active.save(update_fields=["is_active"])
            _sync_live_file(key, new_active)
