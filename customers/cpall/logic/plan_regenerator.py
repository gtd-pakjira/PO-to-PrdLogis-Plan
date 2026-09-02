"""
plan_regenerator.py — Phase 1.6 sub-phase 4: สร้างไฟล์ Excel ใหม่ตอนกด "ดาวน์โหลด" จากข้อมูลดิบ
(plan_sku_result) + เทมเพลตเวอร์ชันที่ผูกไว้กับแผนนั้น (production_template_version / template_version)

*** หลักการสำคัญ: ไม่คำนวณอะไรเองเลย เขียนแค่ "ค่าดิบ" ที่ Admin กรอก/ระบบคำนวณจาก PO เท่านั้น ***
(ยอดสั่งต่อ SKU/คอลัมน์, ยอดเผื่อ) ส่วนสูตรทั้งหมด (ยอดรวม, ยอดคืน, แตกลัง/เศษ) "ปล่อยทิ้งไว้เป็นสูตร"
ในเทมเพลตต่อไป ไม่แตะเลย — ไฟล์ที่ได้จะมีสูตรจริงครบเหมือนที่ Admin คุ้นเคย เปิดด้วย Excel/LibreOffice
จริงแล้วคำนวณเองอัตโนมัติ (ตรงกับที่ยืนยันไว้ว่าต้องมีสูตรอยู่ในไฟล์ที่ดาวน์โหลด)

ทำงานได้เฉพาะแผนที่สร้างหลัง Phase 1.6 sub-phase 3+5 (มีข้อมูลใน plan_sku_result และผูก template
version ไว้แล้ว) — แผนเก่ากว่านั้น raise PlanRegenerateError ให้ผู้เรียก fallback ไปเสิร์ฟไฟล์เดิมที่
ยังเก็บไว้บนดิสก์แทน (ไฟล์เก่ายังไม่ถูกลบในเฟสนี้ — เก็บไว้เป็นทางสำรองก่อน)
"""
import io

import openpyxl

from customers.cpall.logic.date_utils import update_date_headers
from customers.cpall.logic.excel_export import BUFFER_COL, BUFFER_ROW_OFFSET, _find_sub_location_columns
from customers.cpall.logic.excel_export import SHEET_NAME as PP_SHEET_NAME
from customers.cpall.logic.excel_export import _find_sku_header_rows as _find_pp_sku_header_rows
from customers.cpall.logic.grouping import get_dates_by_sub_location
from customers.cpall.logic.logistic_plan_export import (
    GROUP_TEMPLATES,
    _find_column_labels,
    _find_line_no_column,
    _find_qty_column_range,
)
from customers.cpall.logic.logistic_plan_export import _find_sku_header_rows as _find_lp_sku_header_rows


class PlanRegenerateError(Exception):
    pass


def _update_dates(ws, plan_run, col_to_sub_location):
    po_import_ids = list(plan_run.po_imports.values_list("id", flat=True))
    if not po_import_ids:
        return
    dates_by_sub_location = get_dates_by_sub_location(po_import_ids)

    def date_resolver(col):
        sub_loc = col_to_sub_location.get(col)
        if sub_loc is None or sub_loc not in dates_by_sub_location:
            return None
        production_date, po_date = dates_by_sub_location[sub_loc]
        if production_date is None or po_date is None:
            return None
        return production_date, po_date

    update_date_headers(ws, date_resolver)


def regenerate_production_plan_bytes(plan_run_id: int) -> bytes:
    """สร้างไฟล์ Production Plan ใหม่จากข้อมูลดิบ คืนเป็น bytes ตรงๆ (ไม่ผ่านไฟล์ชั่วคราวบนดิสก์เลย)"""
    from customers.cpall.models import PlanRun, PlanSkuResult

    try:
        plan_run = PlanRun.objects.get(id=plan_run_id)
    except PlanRun.DoesNotExist:
        raise PlanRegenerateError("ไม่พบแผนนี้")

    if plan_run.production_template_version is None:
        raise PlanRegenerateError("แผนนี้ไม่ได้ผูกเทมเพลตเวอร์ชันไว้ (อาจเป็นแผนเก่าก่อนมีระบบนี้)")

    results = list(PlanSkuResult.objects.filter(plan_run_id=plan_run_id, sheet_type="production"))
    if not results:
        raise PlanRegenerateError("ไม่มีข้อมูลดิบของแผนนี้ (อาจเป็นแผนเก่าก่อนมีระบบนี้)")

    by_barcode = {}
    for r in results:
        entry = by_barcode.setdefault(r.barcode, {"buffer_qty": r.buffer_qty, "qty_by_col": {}})
        entry["qty_by_col"][r.column_label] = r.qty

    wb = openpyxl.load_workbook(plan_run.production_template_version.file_path)
    ws = wb[PP_SHEET_NAME]

    col_to_sub_location = _find_sub_location_columns(ws)
    sub_location_to_col = {v: k for k, v in col_to_sub_location.items()}
    header_rows = _find_pp_sku_header_rows(ws)

    for barcode, row in header_rows.items():
        data = by_barcode.get(barcode)
        if data is None:
            continue  # SKU ในเทมเพลตแต่ไม่มีข้อมูล (ไม่เคยสั่งรอบนี้) -> ปล่อยว่างไว้เหมือนตอนสร้างแผนจริง
        for sub_loc, qty in data["qty_by_col"].items():
            col = sub_location_to_col.get(sub_loc)
            if col is not None and qty is not None:
                ws.cell(row=row, column=col, value=float(qty))
        if data["buffer_qty"] is not None:
            ws.cell(row=row + BUFFER_ROW_OFFSET, column=BUFFER_COL, value=float(data["buffer_qty"]))

    _update_dates(ws, plan_run, col_to_sub_location)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def regenerate_logistic_plan_bytes(plan_run_id: int, group_name: str) -> bytes:
    """สร้างไฟล์ Logistic Plan ใหม่จากข้อมูลดิบสำหรับ 1 กลุ่ม คืนเป็น bytes ตรงๆ"""
    from customers.cpall.models import PlanRun, PlanRunLogisticFile, PlanSkuResult

    try:
        plan_run = PlanRun.objects.get(id=plan_run_id)
    except PlanRun.DoesNotExist:
        raise PlanRegenerateError("ไม่พบแผนนี้")

    logistic_file = PlanRunLogisticFile.objects.filter(
        plan_run_id=plan_run_id, group_name=group_name
    ).first()
    if logistic_file is None or logistic_file.template_version is None:
        raise PlanRegenerateError("กลุ่มนี้ของแผนไม่ได้ผูกเทมเพลตเวอร์ชันไว้ (อาจเป็นแผนเก่าก่อนมีระบบนี้)")

    results = list(PlanSkuResult.objects.filter(
        plan_run_id=plan_run_id, sheet_type="logistic", group_name=group_name
    ))
    if not results:
        raise PlanRegenerateError("ไม่มีข้อมูลดิบของกลุ่มนี้ (อาจเป็นแผนเก่าก่อนมีระบบนี้)")

    by_barcode = {}
    for r in results:
        entry = by_barcode.setdefault(r.barcode, {})
        entry[r.column_label] = r.qty

    wb = openpyxl.load_workbook(logistic_file.template_version.file_path)
    _, sheet_name = GROUP_TEMPLATES[group_name]
    ws = wb[sheet_name]

    line_no_col, header_row = _find_line_no_column(ws)
    name_col = line_no_col + 1
    qty_start_col = line_no_col + 3
    qty_start_col, qty_end_col = _find_qty_column_range(ws, qty_start_col, header_row)

    col_labels = {}
    last_sub_location = None
    for col in range(qty_start_col, qty_end_col + 1):
        sub_loc, po_idx = _find_column_labels(ws, col, header_row)
        if sub_loc is None:
            sub_loc = last_sub_location if last_sub_location is not None else group_name
        last_sub_location = sub_loc
        col_labels[col] = (sub_loc, po_idx)
    if len(col_labels) == 1:
        only_col = list(col_labels.keys())[0]
        sub_loc, po_idx = col_labels[only_col]
        if po_idx is None:
            col_labels[only_col] = (sub_loc, 1)

    label_to_col = {}
    for col, (sub_loc, po_idx) in col_labels.items():
        label = f"{sub_loc} PO{po_idx}" if po_idx else sub_loc
        label_to_col[label] = col

    header_rows = _find_lp_sku_header_rows(ws, name_col)

    for barcode, row in header_rows.items():
        data = by_barcode.get(barcode)
        if data is None:
            continue
        for label, qty in data.items():
            col = label_to_col.get(label)
            if col is not None and qty is not None:
                ws.cell(row=row, column=col, value=float(qty))

    col_to_sub_location = {col: sub_loc for col, (sub_loc, _) in col_labels.items()}
    _update_dates(ws, plan_run, col_to_sub_location)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
