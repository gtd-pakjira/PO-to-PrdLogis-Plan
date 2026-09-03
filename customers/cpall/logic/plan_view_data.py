"""
plan_view_data.py — อ่านค่ากลับจากไฟล์ Excel ที่ระบบสร้างไว้แล้ว (Production Plan / Logistic Plan)
มาจัดเป็นตาราง (list of dict) สำหรับแสดงในหน้าเว็บ — ไม่คำนวณเลขใหม่เอง แค่ "อ่านของจริงที่มีอยู่แล้ว"
ยกเว้น 2 ค่าที่เป็นสูตร Excel ที่ openpyxl อ่านค่าที่คำนวณแล้วไม่ได้ (ดูหมายเหตุด้านล่าง) — จำลองสูตร
เดียวกันเป๊ะด้วย Python แทน เพื่อให้หน้าเว็บโชว์ "ครบเหมือนไฟล์ Excel จริง" ไม่ใช่แค่บางส่วน

ทำไมอ่านจากไฟล์ที่สร้างแล้ว แทนที่จะคำนวณแยกอีกชุด: กันไม่ให้ตัวเลขในหน้าเว็บกับในไฟล์ที่ดาวน์โหลด
ไปคนละทางกัน (ถ้าคำนวณ 2 ที่แยกกัน มีความเสี่ยงว่าจะแก้โค้ดจุดหนึ่งแล้วลืมอีกจุด ทำให้เลขไม่ตรงกัน)

หมายเหตุเรื่องสูตร Excel: คอลัมน์ "ยอดรวม", ตัวเลขแตกลัง/เศษ (แพค), และ "ยอดคืน" ในไฟล์จริงเป็นสูตร
Excel (=SUM(...), =IF(...)) ซึ่ง openpyxl อ่านค่าที่คำนวณแล้วไม่ได้ (เพราะไฟล์เพิ่งสร้างจาก openpyxl
เอง ไม่เคยผ่านการเปิดด้วย Excel/LibreOffice จริง) — เลยคำนวณเองในโค้ด Python แทน โดยจำลองสูตรเดียวกัน
เป๊ะกับที่อยู่ในไฟล์เทมเพลตจริง (ยืนยันจากการอ่านสูตรตรงๆ จากไฟล์)
"""
import math
import re

import openpyxl

from customers.cpall.logic.excel_export import (
    BUFFER_COL,
    BUFFER_ROW_OFFSET,
    _find_sub_location_columns,
)
from customers.cpall.logic.excel_export import (
    COL_NAME as PP_COL_NAME,
)
from customers.cpall.logic.excel_export import (
    SHEET_NAME as PP_SHEET_NAME,
)
from customers.cpall.logic.excel_export import (
    _find_sku_header_rows as _find_pp_sku_header_rows,
)
from customers.cpall.logic.logistic_plan_export import (
    _find_column_labels,
    _find_line_no_column,
    _find_qty_column_range,
    get_group_templates,
)
from customers.cpall.logic.logistic_plan_export import (
    _find_sku_header_rows as _find_lp_sku_header_rows,
)

PP_COL_PRICE = 5   # E
PP_COL_PACK = 6    # F
PP_NAME_EN_ROW_OFFSET = 1        # แถวชื่ออังกฤษ = แถวหัว SKU + 1
PP_RETURN_ROW_OFFSET = 3         # แถว "ยอดคืน" = แถวหัว SKU + 3
PP_RETURN_COL = BUFFER_COL       # ยอดคืนอยู่คอลัมน์เดียวกับยอดเผื่อ (O) แต่คนละแถว

# กลุ่มจุดส่งที่ใช้คำนวณ "ยอดคืน" (= ยอดเผื่อ - ยอดที่ใช้จริงในกลุ่มนี้) — จำลองจากสูตรจริงในเทมเพลต
# "=SUM(O13-M11-N11-O11-P11-Q11)" ซึ่ง M:Q คือ 5 จุดส่งสุดท้าย (รอบเช้าต่างจังหวัด) เป๊ะ
RETURN_GROUP_SUB_LOCATIONS = ["ขอนแก่น", "สุราษฎร์ธานี", "เชียงใหม่", "ภูเก็ต", "หาดใหญ่"]


def _pack_breakdown_text(qty, pack_size):
    """
    จำลองสูตร Excel: IF(qty=0,"",IF(MOD(qty,pack)=0, qty/pack, IF(INT(qty/pack)=0, MOD&" P", INT&" + "&MOD&" P")))
    แปลงยอดสั่ง (ลัง) เป็นข้อความแตกลัง/เศษ เช่น 153 ลัง บรรจุ 36/ลัง -> "4 + 9 P" (4 ลังเต็ม + 9 ชิ้นเศษ)
    """
    try:
        qty = float(qty or 0)
        pack_size = float(pack_size or 0)
    except (TypeError, ValueError):
        return ""
    if qty == 0 or pack_size == 0:
        return ""
    qty_i, pack_i = int(qty), int(pack_size)
    remainder = qty_i % pack_i
    whole = qty_i // pack_i
    if remainder == 0:
        return str(whole)
    if whole == 0:
        return f"{remainder} P"
    return f"{whole} + {remainder} P"


def _basket_total(qty_by_column: dict, pack_size) -> int:
    """
    จำลองสูตร Excel ในชีต '-รถ'/'มหาชัย' ของแต่ละกลุ่ม: =SUMPRODUCT(ROUNDUP(qty/pack, 0))
    ปัดเศษแต่ละคอลัมน์ (จุดส่ง x PO) ขึ้นเป็นตะกร้าเต็มก่อน แล้วรวมทุกคอลัมน์ — เป็นสูตรง่ายๆ ที่คงที่
    เหมือนกันทุกกลุ่มพื้นที่ (ต่างจากเรื่องจัดรถ/vehicle assignment ในชีต "คันที่ 1/2" ที่ซับซ้อนกว่า
    และยังไม่ได้ทำ เพราะรอข้อมูลเพิ่มเติม — อันนี้แค่ "ตะกร้ารวมกี่ใบ" ไม่เกี่ยวกับเรื่องจัดรถ)
    """
    try:
        pack_size = float(pack_size or 0)
    except (TypeError, ValueError):
        return 0
    if pack_size == 0:
        return 0
    total = 0
    for qty in qty_by_column.values():
        if qty:
            total += math.ceil(qty / pack_size)
    return total


def get_production_plan_table(filepath: str) -> dict:
    """
    -> {
        "sub_locations": ["บางบัวทอง", "ชลบุรี", ...],
        "rows": [
            {"barcode":, "name_th":, "name_en":, "price":, "pack_size":,
             "qty_by_location": {...}, "pack_text_by_location": {...},
             "grand_total":, "buffer_qty":, "return_qty":},
            ...
        ]
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[PP_SHEET_NAME]

    col_to_sub_location = _find_sub_location_columns(ws)
    header_rows = _find_pp_sku_header_rows(ws)  # {barcode: row}, ไม่เรียงลำดับ

    rows = []
    for barcode, row in sorted(header_rows.items(), key=lambda kv: kv[1]):  # เรียงตามแถวในไฟล์
        name_th = ws.cell(row=row, column=PP_COL_NAME).value
        name_en = ws.cell(row=row + PP_NAME_EN_ROW_OFFSET, column=PP_COL_NAME).value
        price = ws.cell(row=row, column=PP_COL_PRICE).value
        pack_size = ws.cell(row=row, column=PP_COL_PACK).value

        qty_by_location, pack_text_by_location = {}, {}
        grand_total = 0
        for col, sub_loc in col_to_sub_location.items():
            val = ws.cell(row=row, column=col).value
            qty_by_location[sub_loc] = val
            pack_text_by_location[sub_loc] = _pack_breakdown_text(val, pack_size)
            if isinstance(val, (int, float)):
                grand_total += val

        buffer_qty = ws.cell(row=row + BUFFER_ROW_OFFSET, column=BUFFER_COL).value
        return_qty = None
        if buffer_qty is not None:
            used_in_return_group = sum(
                (qty_by_location.get(loc) or 0) for loc in RETURN_GROUP_SUB_LOCATIONS
            )
            return_qty = buffer_qty - used_in_return_group

        rows.append({
            "barcode": barcode,
            "name_th": name_th,
            "name_en": name_en,
            "price": price,
            "pack_size": pack_size,
            "qty_by_location": qty_by_location,
            "pack_text_by_location": pack_text_by_location,
            "grand_total": grand_total,
            "buffer_qty": buffer_qty,
            "return_qty": return_qty,
        })

    return {"sub_locations": list(col_to_sub_location.values()), "rows": rows}


def get_logistic_plan_table(filepath: str, group_name: str) -> dict:
    """
    -> {
        "columns": ["บางบัวทอง PO1", "ชลบุรี PO1", "ชลบุรี PO2", ...],
        "rows": [{"barcode":, "name_th":, "name_en":, "pack_size":,
                   "qty_by_column": {...}, "pack_text_by_column": {...}, "grand_total":}, ...]
    }
    """
    _, sheet_name = get_group_templates()[group_name]
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    line_no_col, header_row = _find_line_no_column(ws)
    name_col = line_no_col + 1
    pack_col = line_no_col + 2
    qty_start_col = line_no_col + 3
    qty_start_col, qty_end_col = _find_qty_column_range(ws, qty_start_col, header_row)

    # หา label ต่อคอลัมน์ (เหมือน logic ใน logistic_plan_export.export_logistic_plan)
    col_labels = {}
    last_sub_location = None
    for col in range(qty_start_col, qty_end_col + 1):
        sub_loc, po_idx = _find_column_labels(ws, col, header_row)
        if sub_loc is None:
            sub_loc = last_sub_location if last_sub_location is not None else group_name
        last_sub_location = sub_loc
        col_labels[col] = (sub_loc, po_idx)
    if len(col_labels) == 1:
        only_col = list(col_labels.keys())[0]
        sub_loc, po_idx = col_labels[only_col]
        if po_idx is None:
            col_labels[only_col] = (sub_loc, 1)

    columns = []
    col_to_label = {}
    for col in sorted(col_labels.keys()):
        sub_loc, po_idx = col_labels[col]
        label = f"{sub_loc} PO{po_idx}" if po_idx else sub_loc
        columns.append(label)
        col_to_label[col] = label

    header_rows = _find_lp_sku_header_rows(ws, name_col)

    rows = []
    for barcode, row in sorted(header_rows.items(), key=lambda kv: kv[1]):
        name_th = ws.cell(row=row, column=name_col).value
        # แถวถัดไปมีชื่ออังกฤษปนกับบาร์โค้ดในเซลล์เดียว เช่น "BlackGrapesSeedless150g.(8859388000025)"
        # — ตัดส่วน "(บาร์โค้ด)" ท้ายข้อความออก เหลือแค่ชื่ออังกฤษล้วนๆ
        raw_name_en = ws.cell(row=row + 1, column=name_col).value
        name_en = raw_name_en
        if isinstance(raw_name_en, str):
            name_en = re.sub(r"\(\d{10,16}\)\s*$", "", raw_name_en).strip()
        pack_size = ws.cell(row=row, column=pack_col).value

        qty_by_column, pack_text_by_column = {}, {}
        grand_total = 0
        for col, label in col_to_label.items():
            val = ws.cell(row=row, column=col).value
            qty_by_column[label] = val
            pack_text_by_column[label] = _pack_breakdown_text(val, pack_size)
            if isinstance(val, (int, float)):
                grand_total += val

        rows.append({
            "barcode": barcode,
            "name_th": name_th,
            "name_en": name_en,
            "pack_size": pack_size,
            "qty_by_column": qty_by_column,
            "pack_text_by_column": pack_text_by_column,
            "grand_total": grand_total,
            "basket_total": _basket_total(qty_by_column, pack_size),
        })

    return {"columns": columns, "rows": rows}


def get_production_plan_table_from_db(plan_run_id: int) -> dict:
    """
    เหมือน get_production_plan_table() แต่อ่านจาก plan_sku_result (Phase 1.6 sub-phase 3) แทนการ
    เปิดไฟล์ Excel — ค่าที่ได้มาจาก LibreOffice คำนวณจริงตอนสร้างแผน ไม่ใช่สูตรจำลอง Python
    คืน dict ว่าง (rows=[]) ถ้าไม่มีข้อมูล (แผนเก่าก่อนมีระบบนี้ หรือ extraction ตอนสร้างแผน fail)
    ให้ผู้เรียก fallback ไปใช้ get_production_plan_table() แทนได้
    """
    from customers.cpall.models import PlanSkuResult

    qs = PlanSkuResult.objects.filter(plan_run_id=plan_run_id, sheet_type="production").order_by("id")

    sub_locations = []
    by_barcode = {}
    for r in qs:
        if r.column_label not in sub_locations:
            sub_locations.append(r.column_label)
        if r.barcode not in by_barcode:
            by_barcode[r.barcode] = {
                "barcode": r.barcode, "name_th": r.name_th, "name_en": r.name_en,
                "price": r.price, "pack_size": r.pack_size,
                "qty_by_location": {}, "pack_text_by_location": {},
                "grand_total": r.grand_total, "buffer_qty": r.buffer_qty, "return_qty": r.return_qty,
            }
        by_barcode[r.barcode]["qty_by_location"][r.column_label] = r.qty
        by_barcode[r.barcode]["pack_text_by_location"][r.column_label] = r.pack_text

    return {"sub_locations": sub_locations, "rows": list(by_barcode.values())}


def get_logistic_plan_table_from_db(plan_run_id: int, group_name: str) -> dict:
    """เหมือน get_logistic_plan_table() แต่อ่านจาก plan_sku_result แทนการเปิดไฟล์ Excel"""
    from customers.cpall.models import PlanSkuResult

    qs = PlanSkuResult.objects.filter(
        plan_run_id=plan_run_id, sheet_type="logistic", group_name=group_name
    ).order_by("id")

    columns = []
    by_barcode = {}
    for r in qs:
        if r.column_label not in columns:
            columns.append(r.column_label)
        if r.barcode not in by_barcode:
            by_barcode[r.barcode] = {
                "barcode": r.barcode, "name_th": r.name_th, "name_en": r.name_en,
                "pack_size": r.pack_size,
                "qty_by_column": {}, "pack_text_by_column": {},
                "grand_total": r.grand_total, "basket_total": r.basket_total,
            }
        by_barcode[r.barcode]["qty_by_column"][r.column_label] = r.qty
        by_barcode[r.barcode]["pack_text_by_column"][r.column_label] = r.pack_text

    return {"columns": columns, "rows": list(by_barcode.values())}
